#!/usr/bin/env python3
"""
Negative fund balance history, per docs/V24A_NEGATIVE_BALANCE_FINDING.md.

Reads the State Controller's raw Financial Transactions Report workbooks
and returns, per entity, WHICH of the eight covered fiscal years it filed
a negative total governmental fund balance in.

WHAT THIS PUBLISHES, AND THE ONE THING IT PUBLISHES INSTEAD OF A FIGURE.

The count of years is the whole product. V24a measured that 132 of the
294 affected districts (45%) are negative in exactly ONE year of eight
while only 11 are negative in all eight, and that the steady annual count
— 87, 89, 83, 95, 93, 89, 89, 95 — conceals a largely rotating cast. A
list or a statewide count presents a one-year event and an eight-year
condition as the same fact. The year count is the only thing in the data
that separates them, so it is what ships, on the entity's own record.

NO DOLLAR FIGURE IS EMITTED, deliberately, and the payload has no field
for one. A deficit figure beside the spending figure already on the same
record invites division — "3% underwater" — which is the ratio V24 §4
refused after measuring it four ways. The absence is structural rather
than editorial: there is nowhere to put a number.

THE LIMITS TRAVEL WITH THE FACT, not in a method note:

  - AS FILED AND UNVERIFIABLE. No SCO control exists (V24 §2.3: a regex
    over all 170 Socrata view names returns zero hits). Of the three
    checks available, two are the Controller restating one figure, and
    the only one involving an independent quantity — beginning + net
    change = end of year — holds for 70.2% of negative-balance entities
    against 84.8% of positive ones. Filing quality is measurably worse
    on precisely the entities being named, and the record says so.
  - NO CAUSE IS PUBLISHED. The filing carries no cause field, note or
    flag, so the data cannot distinguish fiscal distress from an
    accounting-timing artifact. Stated plainly; never characterised.
  - THE WINDOW EDGE IS A REAL LIMIT. A single negative year at either end
    of the window cannot be told from a longer run that continues outside
    it. 75 of the 132 single-year cases are in that position, and the
    affected records say so rather than implying a one-year event.

Requires openpyxl.
"""

import io
import json
import re
import sys
import urllib.request
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import cache_guard                               # noqa: E402

CACHE = Path(__file__).resolve().parent / "cache" / "ftr"
UA = {"User-Agent": "Citizen Ledger data pipeline"}

# DECLARED PER VINTAGE, never discovered. Each workbook covers the fiscal
# years named here and no others; a vintage that turns out to carry a
# different set stops the build rather than being absorbed silently.
SOURCES = {
    "district": [
        ("uvvz-zjub", ("2017",)),
        ("3qib-6kft", ("2018", "2019")),
        ("c2qj-ad4e", ("2020",)),
        ("ftzy-54cr", ("2021",)),
        ("ffcm-gghd", ("2022",)),
        ("dp5e-7wm8", ("2023", "2024")),
    ],
    "city": [
        ("kcfp-hz7x", ("2017",)),
        ("885w-tc2s", ("2018", "2019", "2020")),
        ("kyrq-f99p", ("2021",)),
        ("vbm4-7c9z", ("2022",)),
        ("wjvf-fpdc", ("2023", "2024")),
    ],
}

# The eight-year window this finding measured. Declared so a workbook
# that starts publishing a ninth year does not silently change the
# denominator in "N of 8" without anyone deciding to.
YEARS = ("2017", "2018", "2019", "2020", "2021", "2022", "2023", "2024")

# The sheet carrying the governmental balance sheet. SCO numbers it
# differently per entity type and per vintage ("22 CIX_BAL_GOVT_FUNDS",
# "37 CIX_BAL_GOVT_FUND", "CIX_BAL_GOVT_FUND"), so it is matched on shape
# rather than name — but the match must be UNIQUE or the build stops.
SHEET_RE = re.compile(r"BAL_GOVT_FUNDS?$", re.I)

# The column is likewise named slightly differently across vintages. Both
# forms end in the same two facts: it is a Total, and it is the Total
# Governmental Funds column.
COL_PREFIX = "Total Fund Balances (Deficits)_"
COL_MARK = "Total Governmental Funds"


def _get(url):
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=900) as r:
        return r.read()


def workbook(view_id, refresh=False):
    """Fetch a raw FTR workbook, cached and read-only.

    EXISTENCE IS CONTENT: the blob is accepted only on the PK magic bytes
    of a zip container, never on a status code. SCO's own /files/latest
    path returns a JSON error with HTTP 200, which is exactly the
    soft-404 this check exists for.
    """
    path = CACHE / f"{view_id}.xlsx"
    if path.exists() and not refresh:
        return path
    meta = json.loads(_get(f"https://bythenumbers.sco.ca.gov/api/views/{view_id}.json"))
    blob = meta.get("blobId")
    if not blob:
        raise SystemExit(f"{view_id}: view metadata carries no blobId — "
                         "the view is not a file view; nothing written")
    body = _get(f"https://bythenumbers.sco.ca.gov/api/views/{view_id}/files/{blob}")
    if body[:2] != b"PK":
        raise SystemExit(
            f"{view_id}: downloaded {len(body)} bytes whose magic is "
            f"{body[:2]!r}, not a zip container. SCO serves a JSON error "
            "with HTTP 200 for paths that do not exist, so this is checked "
            "on content; nothing written")
    CACHE.mkdir(parents=True, exist_ok=True)
    cache_guard.write_cached(path, body, binary=True)
    return path


def _sheet(wb, view_id):
    hits = [s for s in wb.sheetnames if SHEET_RE.search(s.strip())]
    if len(hits) != 1:
        raise SystemExit(
            f"{view_id}: expected exactly one governmental balance-sheet "
            f"tab, found {hits or 'none'}. The workbook shape changed and "
            "reading the wrong tab would publish the wrong entities; "
            "nothing written")
    return wb[hits[0]]


def _key_columns(header, view_id):
    """Locate Entity ID / Entity Name / Fiscal Year BY HEADER NAME.

    Reading them positionally is the defect this function replaces. The
    FY2022-23 workbooks put them at 0/1/2; the FY2016-17 city workbook
    does not, and a positional read there returned entity ids where the
    fiscal year was expected — a wrong column that produced confident
    values rather than an error, which is the failure StrictRow exists
    to make loud. Names differ slightly across vintages, so each is
    matched on a normalised form and must resolve to exactly one column.
    """
    return (_find_one(header, view_id, "Entity ID", "EntityID"),
            _find_one(header, view_id, "Entity Name", "EntityName"),
            _find_one(header, view_id, "Fiscal Year", "FiscalYear"))


def _find_one(header, view_id, *aliases):
    want = {a.replace(" ", "").lower() for a in aliases}
    hits = [i for i, h in enumerate(header)
            if h and str(h).replace(" ", "").lower() in want]
    if len(hits) != 1:
        raise SystemExit(
            f"{view_id}: expected exactly one of {sorted(want)} among the "
            f"columns, found {len(hits)}; nothing written")
    return hits[0]


def _column(header, view_id):
    hits = [i for i, h in enumerate(header)
            if h and str(h).startswith(COL_PREFIX) and COL_MARK in str(h)]
    if len(hits) != 1:
        raise SystemExit(
            f"{view_id}: expected exactly one "
            f"'{COL_PREFIX}…{COL_MARK}' column, found {len(hits)}; "
            "nothing written")
    return hits[0]


def _counties(wb, view_id):
    """Entity ID -> county, from the workbook's own ENTITIES sheet.

    The site's district roster is keyed on (name, county) because the
    Socrata feed exposes no entity id — so joining this history to it
    needs the county, and taking it from the same workbook keeps the pair
    internally consistent rather than inferred.
    """
    hits = [s for s in wb.sheetnames if s.strip().upper().endswith("ENTITIES")]
    if len(hits) != 1:
        raise SystemExit(f"{view_id}: expected one ENTITIES tab, found "
                         f"{hits or 'none'}; nothing written")
    ws = wb[hits[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    # only the id and the county are needed here; a single-year workbook's
    # ENTITIES tab carries no Fiscal Year column, so requiring all three
    # would refuse a vintage that is perfectly well formed
    c_id = _find_one(header, view_id, "Entity ID", "EntityID")
    cty = [i for i, h in enumerate(header)
           if h and str(h).replace(" ", "").lower() in ("countyname", "county")]
    if len(cty) != 1:
        raise SystemExit(f"{view_id}: expected one county column on the "
                         f"ENTITIES tab, found {len(cty)}; nothing written")
    return {str(r[c_id]).strip(): str(r[cty[0]] or "").strip() for r in rows}


def history(kind, refresh=False):
    """(entity id) -> {"name": str, "years": [fy, ...]} for entities that
    filed a NEGATIVE total governmental fund balance in at least one year.

    Keyed on the Controller's own Entity ID, never on the name. The name
    is carried for display only. Two governments sharing a name is a
    documented hazard on this site (Rural North Vacaville), and a
    name-keyed history would merge their filings.
    """
    import openpyxl
    seen_years, out = set(), {}
    for view_id, declared in SOURCES[kind]:
        path = workbook(view_id, refresh)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        cty = _counties(wb, view_id)
        ws = _sheet(wb, view_id)
        rows = ws.iter_rows(values_only=True)
        header = list(next(rows))
        col = _column(header, view_id)
        c_id, c_name, c_fy = _key_columns(header, view_id)
        found = set()
        for r in rows:
            fy = str(r[c_fy]).strip()
            found.add(fy)
            v = r[col]
            if not isinstance(v, (int, float)):
                continue        # a non-numeric cell is not-published, not zero
            if v >= 0:
                continue
            eid = str(r[c_id]).strip()
            rec = out.setdefault(eid, {"name": str(r[c_name]).strip(),
                                       "county": cty.get(eid, ""), "years": []})
            if not rec.get("county"):
                rec["county"] = cty.get(eid, "")
            if fy not in rec["years"]:
                rec["years"].append(fy)
        wb.close()
        if found != set(declared):
            raise SystemExit(
                f"{view_id}: declared fiscal years {sorted(declared)} but the "
                f"workbook carries {sorted(found)}. A vintage that changes "
                "its coverage changes the denominator in 'N of 8' without "
                "anyone deciding to; nothing written")
        seen_years |= found
    if seen_years != set(YEARS):
        raise SystemExit(
            f"{kind}: the declared vintages cover {sorted(seen_years)}, not "
            f"the {len(YEARS)}-year window {list(YEARS)}; nothing written")
    for rec in out.values():
        rec["years"].sort()
    return out


def edge_cases(years):
    """True when the entity's negative years touch either end of the window.

    A single negative year at an edge cannot be told from a longer run
    continuing outside the window — measured, 75 of 132 single-year cases.
    The affected records say so rather than implying a one-year event.
    """
    return bool(years) and (YEARS[0] in years or YEARS[-1] in years)


def main():
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--kind", choices=sorted(SOURCES), default="district")
    ap.add_argument("--refresh", action="store_true")
    a = ap.parse_args()
    h = history(a.kind, a.refresh)
    n1 = sum(1 for r in h.values() if len(r["years"]) == 1)
    nall = sum(1 for r in h.values() if len(r["years"]) == len(YEARS))
    print(f"{a.kind}: {len(h)} entities negative in at least one of "
          f"{len(YEARS)} years; {n1} in exactly one; {nall} in all",
          file=sys.stderr)
    for fy in YEARS:
        print(f"   FY{fy}: {sum(1 for r in h.values() if fy in r['years'])}",
              file=sys.stderr)


if __name__ == "__main__":
    main()
