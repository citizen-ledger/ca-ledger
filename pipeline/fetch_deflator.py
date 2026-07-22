#!/usr/bin/env python3
"""
Citizen Ledger — the price deflator (V14).

Rebuilds ../deflator-data.js from the California Department of Finance's
published fiscal-year deflator file.

WHAT THIS IS, AND WHAT IT IS NOT — read this before changing anything:

  Every other figure on this site is REPRODUCTION. A city's expenditure
  is the Controller's figure; a district's Current Expense of Education
  is CDE's figure, reproduced to the cent. This file is different, and
  the difference is stated on the face of the feature rather than
  smoothed over.

  NO CALIFORNIA SOURCE DEFLATES ITS OWN PUBLISHED SERIES. Verified in
  docs/V14_INFLATION_FINDING.md: DOF publishes a fifty-year spending
  series and never deflates it; CDE's Current Expense workbook — the
  exact artifact this site consumes — has no constant-dollar column;
  the State Controller is an explicit pass-through of as-submitted
  figures. There is no official practice to adopt. The inflation
  adjustment is THE LEDGER'S OWN METHODOLOGICAL CHOICE, and the pages
  say so in those words.

WHY THIS INDEX. California statute names exactly one price deflator for
government spending: Education Code 42238.1(a)(2), the K-12 COLA, which
specifies

    "the Implicit Price Deflator for State and Local Government
     Purchases of Goods and Services for the United States, as published
     by the United States Department of Commerce ... as reported by the
     Department of Finance."

So the index is the one California law names, and DOF — the department
that statute designates as its reporter — publishes it. Two honest
limits travel with it, and both appear on the page: the index is
NATIONAL, not Californian (DOF states deflators are not available below
the national level), and the LAO, which used this index explicitly from
1999 to 2008, called it "not a particularly good indicator of increases
in school costs."

WHY DOF'S FILE RATHER THAN OUR OWN AVERAGING. Ledger data is fiscal
(1 July - 30 June); price indices are monthly or calendar-annual.
Choosing among a twelve-month mean, a midpoint month, or the calendar
year the fiscal year begins or ends is exactly the kind of unforced
decision this project avoids. DOF publishes the index ALREADY AVERAGED
TO CALIFORNIA FISCAL YEARS, so we adopt that file wholesale and convert
nothing. Do not add an averaging step here.

FORECAST YEARS ARE NOT DEFLATED. DOF flags 2025-26 onward as forecast
(`f/`) — and 2025-26 is the state layer's newest year. A real figure
resting on a projected deflator would change when DOF revises while the
nominal figure never moved. Those years ship with their index value and
an explicit `forecast` flag, and the pages refuse to adjust them.

VINTAGE. A fixed base year does not mean fixed values: BEA revises the
index annually and DOF republishes each May and November. The file's own
"Updated:" stamp, the source URL, and the SHA-256 of the bytes we parsed
are all recorded in meta, so a real figure can always be traced to the
exact index vintage that produced it — and a republication shows up in
the change record like any other moved figure.

Source (public, keyless, no bot-gate — verified):
  https://dof.ca.gov/media/docs/forecasting/economics/economic-indicators/
      inflation/Implicit-Price-Deflators-FY.xlsx

Usage:
    python3 fetch_deflator.py            # dry run
    python3 fetch_deflator.py --write
    python3 fetch_deflator.py --write --refresh   # re-download
Requires openpyxl (already a pipeline dependency).
"""

import argparse
import hashlib
import json
import re
import sys
import urllib.request
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import gates                                     # noqa: E402
from integrity import stamp  # noqa: E402
import revisions  # noqa: E402

import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
CACHE = Path(__file__).resolve().parent / "cache" / "deflator"
OUT_PATH = ROOT / "deflator-data.js"

DOF_BASE = ("https://dof.ca.gov/media/docs/forecasting/economics/"
            "economic-indicators/inflation/")
IPD_FILE = "Implicit-Price-Deflators-FY.xlsx"
IPD_URL = DOF_BASE + IPD_FILE

# The column DOF publishes for the Education Code 42238.1 index. Named
# rather than positional: DOF also publishes GDP and PCE columns in the
# same sheet, and picking the wrong one would be silent.
COLUMN = "State and Local Index"
SHEET = "Deflators_FY"

STATUTE = ("California Education Code 42238.1(a)(2) — the K-12 statutory "
           "cost-of-living index")
INDEX_NAME = ("Implicit Price Deflator for State and Local Government "
              "Purchases of Goods and Services")


def fetch(refresh=False):
    CACHE.mkdir(parents=True, exist_ok=True)
    dest = CACHE / IPD_FILE
    if dest.exists() and not refresh:
        return dest.read_bytes()
    print(f"  downloading {IPD_URL}", file=sys.stderr)
    req = urllib.request.Request(
        IPD_URL, headers={"User-Agent": "ca-ledger-pipeline/1.0"})
    blob = urllib.request.urlopen(req, timeout=120).read()
    if blob[:2] != b"PK":                       # xlsx is a zip
        raise SystemExit("deflator: response is not an .xlsx — refusing to "
                         "parse. Nothing written.")
    dest.write_bytes(blob)
    return blob


def parse(blob):
    dest = CACHE / IPD_FILE
    wb = openpyxl.load_workbook(dest, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"deflator: sheet {SHEET!r} is gone from DOF's file "
                         f"(found {wb.sheetnames}) — the format changed. "
                         "Nothing written; a human must look.")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))

    header = [str(c or "") for c in rows[1]]
    try:
        col = next(i for i, h in enumerate(header) if COLUMN in h)
    except StopIteration:
        raise SystemExit(f"deflator: column {COLUMN!r} is gone from DOF's "
                         f"file (found {header}) — nothing written.")

    title = str(rows[0][0] or "")
    m = re.search(r"\((\d{4})=100\)", title)
    if not m:
        raise SystemExit(f"deflator: cannot read the index base from the "
                         f"sheet title {title!r} — nothing written.")
    index_base = m.group(1)

    series, forecast, notes = {}, set(), []
    seen, dupes = {}, {}
    for r in rows:
        cell = str(r[0] or "").strip()
        if not cell:
            continue
        if not re.match(r"^\d{4}-\d{2}", cell):
            if any(k in cell for k in ("Updated:", "Source:", "Note:",
                                       "f/", "Next forecast")):
                notes.append(cell)
            continue
        fy = cell[:7]
        try:
            val = float(r[col])
        except (TypeError, ValueError):
            continue
        # DOF's file has repeated a fiscal year with conflicting values
        # before (two 2029-30 rows, May 2026 vintage). Taking the last
        # silently would be a coin toss dressed as a figure.
        if fy in seen and abs(seen[fy] - val) > 1e-9:
            dupes.setdefault(fy, [seen[fy]]).append(val)
        seen[fy] = val
        series[fy] = val
        if "f/" in cell:
            forecast.add(fy)

    # A duplicated ACTUAL year is unresolvable and blocks the build. A
    # duplicated FORECAST year is dropped and named — forecast years are
    # never used to adjust anything, so the anomaly is recorded rather
    # than allowed to stop the site. Same discipline as the state layer's
    # named SOURCE_RESIDUAL: the anomaly is surfaced, not absorbed.
    anomalies = []
    for fy, vals in sorted(dupes.items()):
        if fy not in forecast:
            raise SystemExit(
                f"deflator: DOF's file gives fiscal year {fy} more than once "
                f"with different values {vals} and it is not a forecast year. "
                "Unresolvable; nothing written.")
        anomalies.append(
            f"DOF's file lists {fy} twice with different values "
            f"({', '.join(f'{v:.5f}' for v in vals)}). It is a forecast year, "
            "which this site never uses to adjust a figure, so it is dropped "
            "rather than guessed at.")
        series.pop(fy, None)
        forecast.discard(fy)

    gates.require_rows(len(series), 60, "deflator fiscal years parsed",
                       "DOF's file is much longer than that.")

    actuals = [fy for fy in series if fy not in forecast]
    last_actual = max(actuals)
    if forecast and min(forecast) <= last_actual:
        raise SystemExit("deflator: a forecast year sorts at or before the "
                         "last actual — the f/ flags are not what we assume. "
                         "Nothing written.")

    vintage = next((n.split("Updated:")[1].strip().rstrip(".")
                    for n in notes if "Updated:" in n), None)
    if not vintage:
        raise SystemExit("deflator: DOF's 'Updated:' stamp is missing — the "
                         "vintage cannot be recorded, so nothing is written.")

    return {"series": series, "forecast": sorted(forecast),
            "anomalies": anomalies,
            "last_actual": last_actual, "index_base": index_base,
            "vintage": vintage, "notes": notes}


def build(refresh=False):
    blob = fetch(refresh)
    p = parse(blob)
    digest = hashlib.sha256(blob).hexdigest()

    # Base year: the latest COMPLETE ACTUAL fiscal year. Fixed, not
    # tracking-latest — a base that moved every year would restate every
    # published real figure annually for no analytical reason.
    base = p["last_actual"]

    payload = {
        "meta": {
            "source": "dof.ca.gov",
            "sourceLabel": ("California Department of Finance — national "
                            "implicit price deflators, California fiscal-year "
                            "averages"),
            "sourceFile": IPD_FILE,
            "sourceUrl": IPD_URL,
            "generated": date.today().isoformat(),
            "index": INDEX_NAME,
            "column": COLUMN,
            "geography": "United States (national)",
            "statute": STATUTE,
            "indexBase": f"{p['index_base']}=100",
            "baseYear": base,
            "lastActual": p["last_actual"],
            "forecastYears": p["forecast"],
            "vintage": p["vintage"],
            "sourceDigest": "sha256:" + digest,
            "dofNotes": p["notes"],
            "sourceAnomalies": p["anomalies"],
            # The sentences the pages render. Kept here so the claim and
            # the data ship together and cannot drift apart.
            "ours": ("This adjustment is the Ledger's, not the source's. "
                     "The Department of Finance, the Department of Education "
                     "and the State Controller each publish these figures in "
                     "nominal dollars only and deflate nothing. The Ledger "
                     "applies the one price deflator California statute names "
                     "for government spending, unchanged, in the fiscal-year "
                     "form DOF publishes."),
            "limits": ("The index is national, not Californian — DOF states "
                       "deflators are not available below the national level. "
                       "The Legislative Analyst's Office, which used this "
                       "index explicitly from 1999 to 2008, has called it "
                       "“not a particularly good indicator of increases "
                       "in school costs.”"),
            # Per-layer, because the sensitivity genuinely differs by window
            # length and a shared sentence would understate it on K-12 and
            # overstate it on the local layers. Measured in
            # docs/V14_INFLATION_FINDING.md.
            "windowNotes": {
                "local": ("Over this layer's eight-year window the choice of "
                          "index barely matters: this deflator and DOF's "
                          "California CPI differ by 0.6 percentage points in "
                          "total, and they disagree about the direction of "
                          "exactly one city in 482. The nominal-to-real "
                          "difference, by contrast, is large — 71 of 482 "
                          "cities rise in nominal dollars and fall in real "
                          "ones over this window, and none go the other way."),
                "k12": ("This layer's window is only three years, and short "
                        "windows are the sensitive case. Over it, this "
                        "deflator and DOF's California CPI differ by 2.68 "
                        "percentage points — roughly 42% of the whole "
                        "measured inflation — which is enough for the choice "
                        "of index to change the SIGN of a real trend. Treat a "
                        "small real change here as indistinguishable from no "
                        "change."),
                "state": ("Over this layer's six-year window the choice of "
                          "index is a minor effect next to the "
                          "nominal-to-real difference, but the newest year "
                          "cannot be adjusted at all — see the forecast note "
                          "above."),
            },
            "shortWindow": ("Short windows are more sensitive to the choice "
                            "of index than long ones. Over the eight-year "
                            "local window this index and DOF's California CPI "
                            "differ by 0.6 percentage points and disagree "
                            "about the direction of one city in 482; over the "
                            "three-year K-12 window they differ by 2.68 "
                            "points, roughly 42% of the measured inflation, "
                            "where the choice can change a trend's sign."),
            "forecastRule": ("Fiscal years DOF flags as forecast are never "
                             "adjusted. A real figure resting on a projected "
                             "deflator would move when DOF revises even though "
                             "the nominal figure never changed."),
            "fixedBaseNote": ("A fixed base year does not mean fixed values: "
                              "BEA revises the index annually and DOF "
                              "republishes it. The vintage above identifies "
                              "the exact index these figures came from, and a "
                              "republication appears in the record of changes "
                              "like any other moved figure."),
        },
        "fy": {k: p["series"][k] for k in sorted(p["series"])},
    }
    stamp(payload)
    return payload


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--refresh", action="store_true",
                    help="re-download DOF's file instead of using the cache")
    args = ap.parse_args()

    payload = build(args.refresh)
    m = payload["meta"]
    print(f"{m['index']}")
    print(f"  statute      {m['statute']}")
    print(f"  publisher    {m['sourceLabel']}")
    print(f"  vintage      {m['vintage']}   ({m['sourceDigest'][:23]}…)")
    print(f"  index base   {m['indexBase']}")
    print(f"  fiscal years {min(payload['fy'])} .. {max(payload['fy'])}  "
          f"({len(payload['fy'])})")
    print(f"  last actual  {m['lastActual']}   base year {m['baseYear']}")
    print(f"  forecast     {', '.join(m['forecastYears'])}  (never adjusted)")
    b = payload["fy"][m["baseYear"]]
    print("\n  restatement factors into "
          f"{m['baseYear']} dollars (sample):")
    for fy in ("2016-17", "2020-21", "2022-23", m["lastActual"]):
        if fy in payload["fy"]:
            print(f"    {fy}  x{b / payload['fy'][fy]:.4f}")
    if not args.write:
        print("\nDry run — nothing written. Use --write.")
        return
    prev = revisions.previous_payload(OUT_PATH)
    body = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    header = ("/* GENERATED by pipeline/fetch_deflator.py on "
              f"{date.today().isoformat()} from the California Department of "
              "Finance's published fiscal-year deflator file. THE INFLATION "
              "ADJUSTMENT IS THE LEDGER'S OWN — DOF, CDE and the State "
              "Controller publish nominal figures only. Do not edit by hand. */\n")
    OUT_PATH.write_text(header + "window.CA_LEDGER_DEFLATOR = " + body + ";\n",
                        encoding="utf-8")
    revisions.record_revision("deflator", prev, payload)
    print(f"\nWrote {OUT_PATH} ({OUT_PATH.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
