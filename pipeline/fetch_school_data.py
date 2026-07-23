#!/usr/bin/env python3
"""
Citizen Ledger — K-12 schools pipeline (V7, option (a): gated and
comparable, comparison scoped to school districts).

Rebuilds ../school-data.js from the California Department of
Education's SACS unaudited-actuals archives.

Sources per fiscal year (verified in docs/V7_SCHOOLS_FINDING.md):
  https://www3.cde.ca.gov/fiscal-downloads/sacs_data/{YYYY-YY}/sacs{yy}.exe
      → sacs{yy}.mdb  (JET4; UserGL general ledger, UserGL_Totals =
        CDE's own county/state rollups, LEAs, Charters, code tables)
  https://www3.cde.ca.gov/fiscal-downloads/charterdata/{YYYY-YY}/alt{yy}data.exe
      → alt{yy}data.mdb (charter Alternative Form + its own totals)
  https://www.cde.ca.gov/ds/fd/ec/documents/currentexpense{yy}.xlsx
      → CDE-published per-district Current Expense of Education
        (EDP 365) with ADA — THE GATE TARGET
  https://www.cde.ca.gov/fg/aa/pa/documents/lcffsummary{yy}.xlsx
      → LCFF entitlement / local revenue (basic-aid flag), COE funded
        ADA

THE GATE (no write on failure, and it is exact):
  1. Every published district's EDP 365 must be REPRODUCED FROM THE
     RAW GENERAL LEDGER to within $0.05 (the investigation reproduced
     932/932 to the cent for FY 2024-25; one $0.02 publication
     rounding artifact exists in FY 2023-24). The formula is CDE's:
     Fund 01; objects 1000-5999, 6500, 7300-7399; minus goals
     7100-7199 and 8100, functions 3700 and 8500, objects 3701-3702;
     district-entity rows only (SchoolCode '0000000').
  2. Aggregating raw UserGL must reproduce CDE's own UserGL_Totals —
     every state cell and every county cell (fund × object) — and
     the Alternative Form data must reproduce its own totals table.
  3. Structural: 58 COEs; the published district count; ADA present.

TIERS ENCODED (per the approved finding):
  - districts: gated per-LEA against a published figure; compared
    per-ADA with three data-derived daggers (basic aid, small-ADA
    funding floors, sponsored-charter distortion) plus the
    commingled-charter limit stated on affected records.
  - county offices: records only, never compared (CDE itself excludes
    them from its per-ADA statistic).
  - charters that file separately (own SACS or Alternative Form):
    records only, gated through the rollups.
  - charters reported inside an authorizer's books: named directory
    entries pointing at the authorizer; the Fund 01 commingled subset
    cannot be separated and the affected district records say so.
  - JPAs (ROPs/SELPAs): out of scope for records in this version;
    their dollars are inside the rollup gates.

THE OVERLAP BLOCK (meta.overlap) is computed here on every run —
statewide LEA revenues by source from the ledger, against the state
layer's K-12 agency read from ../data.js — so the UI's
figures-do-not-add statement renders live values and hardcodes none.

Usage:
    python3 fetch_school_data.py            # dry run: fetch + gates
    python3 fetch_school_data.py --write    # rebuild ../school-data.js
Requires: mdbtools (mdb-export) and openpyxl — pipeline-only deps,
like pypdf for the state actuals.
"""

import argparse
import csv
import json
import re
import subprocess
import sys
import urllib.request
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import gates  # noqa: E402
import cache_guard                              # noqa: E402
from integrity import stamp  # noqa: E402
import revisions  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
CACHE = Path(__file__).resolve().parent / "cache" / "schools"
OUT_PATH = ROOT / "school-data.js"

# NCES LEA ids for the address view's identifier-based matching: the
# Census geocoder's school-district GEOID equals the NCES LEA id, and
# CDE's public directory carries NCESDist per district. Five SACS
# filers are common-administration arrangements (two legal districts
# filing one report, per CDE's readme); the directory lists their
# CONSTITUENT districts, so those constituents' NCES ids map to the
# filer. This constant is the documented administrative structure,
# verified against the directory — not name matching.
PUBSCHLS = CACHE / "pubschls.txt"   # CDE directory export (live URL is
# bot-gated; obtainable via web.archive.org snapshot of
# https://www.cde.ca.gov/schooldirectory/report?rid=dl1&tp=txt)
# ── RE-CODED DISTRICTS: one body, one NCES id, two California keys ───
#
# California's (Ccode, Dcode) key is administrative and CAN MOVE. The
# NCES id does not. Where SACS files a district under one key and later
# another, the two keys are the SAME DISTRICT and its series is
# continuous; treating them as two records would publish one body twice,
# with a short series each, as if they were different districts.
#
# THE CROSSWALK CANNOT SEE THIS. nces_ids is built from ONE
# current-vintage directory export, so it resolves only the current key
# and any earlier key looks like a district with no NCES id. The next
# re-coding will look exactly like this one — see docs/OPEN.md.
#
# DECLARED, NEVER INFERRED, AND EARNED: assert_recodings() below requires
# BOTH keys to actually occur in the loaded years, so an entry cannot
# rot into a stale exemption once a window no longer reaches back to it.
#
# Lowell Joint Elementary — verified four ways before declaring
# (docs/V16_NCES_IDENTIFIER_FINDING.md):
#   1. the keys NEVER co-occur: 19/64766 in FY2016-17..FY2020-21,
#      30/64766 from FY2021-22, one row each year, never both;
#   2. the boundary is not a step change: +14.7% across it, the 70th
#      percentile of 973 districts in a year whose median was +10.9%
#      (the first full year of federal COVID relief);
#   3. NO other district holds either key in any of the nine years, and
#      the name and type are identical throughout;
#   4. CDE's directory carries NCES 0623010 as ONE Active district row,
#      with no ClosedDate — not a closure and a new opening.
# It is a joint district straddling the Los Angeles / Orange county
# line; CDE moved its administrative county from 19 to 30.
RECODED_DISTRICTS = {
    ("19", "64766"): {
        "to": ("30", "64766"),
        "nces": "0623010",
        "changedAt": "2021-22",
        "why": "Lowell Joint is a joint elementary district straddling the "
               "Los Angeles / Orange county line. California moved its "
               "administrative county code from 19 to 30 at FY2021-22; its "
               "federal NCES identifier, 0623010, did not change. The nine "
               "years are one continuous district.",
    },
}


def assert_recodings(seen_keys):
    """A DECLARED RE-CODING MUST BE EXERCISED BY THE DATA.

    `seen_keys` is every (Ccode, Dcode) the loaded years actually
    contain. Both sides of a declared re-coding must appear, or the
    declaration is making a claim about years this build cannot see —
    which is how an exemption goes stale and starts hiding a real gap.
    """
    for old, rec in RECODED_DISTRICTS.items():
        new = rec["to"]
        if old not in seen_keys or new not in seen_keys:
            raise SystemExit(
                f"DECLARED RE-CODING NOT EXERCISED: {old} -> {new} is "
                f"declared in RECODED_DISTRICTS, but "
                f"{'the old' if old not in seen_keys else 'the new'} key does "
                "not occur in the loaded years. A declaration that nothing "
                "tests is a stale exemption; remove it or widen the window. "
                "Nothing written.")


COMMON_ADMIN_NCES = {
    "0603090": "2376349", "0631230": "2376349",   # Arena Union / Point Arena
    "0625130": "5040717", "0625150": "5040717",   # Modesto City Elem / High
    "0630230": "4940246", "0630250": "4940246",   # Petaluma Elem / JUH
    "0635590": "4440261", "0635600": "4440261",   # Santa Cruz City Elem / High
    "0635810": "4940253", "0635830": "4940253",   # Santa Rosa Elem / High
}

# ---------------------------------------------------------------- vintages
#
# CDE'S SOURCE VOCABULARY, DECLARED PER YEAR — never sniffed.
#
# The Current Expense workbook and the SACS database rename things between
# vintages. A detector loosened to accept any of several spellings would
# also silently accept a vintage nobody has checked, which is how FY2018-19
# used to pass every gate on ZERO districts: its header row reads
# "CO Code"/"District Code" where the shipped detector tested for "CO"
# exactly, so the control table stayed empty and each gate iterated nothing.
#
# Every entry below was read off the real published file, not inferred:
#   sheet / county-code / district-code / district-name column headers.
CE_VINTAGE = {
    "1617": ("CDS",               "CO",      "CDS",           "DISTRICT"),
    "1718": ("CDS",               "CO",      "CDS",           "DISTRICT"),
    "1819": ("District (by CDS)", "CO Code", "District Code", "District"),
    "1920": ("CDS",               "CO",      "CDS",           "DISTRICT"),
    "2021": ("District",          "CO",      "CDS",           "DISTRICT"),
    "2122": ("District",          "CO",      "CDS",           "DISTRICT"),
    "2223": ("District",          "CO",      "CDS",           "District"),
    "2324": ("District",          "CO",      "CDS",           "District"),
    "2425": ("District",          "CO",      "CDS",           "District"),
}

# WHAT THE LCFF SUMMARY PUBLISHES, PER YEAR — declared, never sniffed.
#
# Verified by fetching every candidate year: CDE publishes no LCFF summary
# at all for FY2016-17, FY2017-18 or FY2018-19 (404), and the FY2019-20
# workbook exists but carries 12 columns with no Local Revenue column,
# where FY2020-21 onward carry 20 including "Total Local Revenue or
# In-Lieu of Property Taxes". Basic aid is derived from local revenue
# against entitlement, so FY2019-20 can give funded ADA but not basic aid.
#
# The unknown set therefore DIFFERS PER FACT, which is why this is a table
# of facts and not a list of bad years.
# THE ALTERNATIVE FORM'S SCHOOL COLUMN, declared per vintage.
#
# Measured across all nine years: the charter Alternative Form calls it
# SchoolID through FY2021-22 and SchoolCode from FY2022-23 — the same
# boundary as the SACS rename already recorded in SACS_ERA. Reading it
# with a "try both" fallback would hide the next rename; the vintage is
# declared, and a year outside the table refuses rather than guessing.
ALT_SCHOOL_COL = {
    "1617": "SchoolID", "1718": "SchoolID", "1819": "SchoolID",
    "1920": "SchoolID", "2021": "SchoolID", "2122": "SchoolID",
    "2223": "SchoolCode", "2324": "SchoolCode", "2425": "SchoolCode",
}


def alt_school_col(yy):
    """The Alternative Form's school-code column for this vintage."""
    if yy not in ALT_SCHOOL_COL:
        raise SystemExit(
            f"FY {yy}: the Alternative Form's school column is not declared "
            "for this vintage. CDE renamed it SchoolID -> SchoolCode at "
            "FY2022-23; add this year to ALT_SCHOOL_COL deliberately after "
            "checking the file. Nothing written.")
    return ALT_SCHOOL_COL[yy]


LCFF_PUBLISHES = {
    "1617": {"basicAid": False, "fundedADA": False},
    "1718": {"basicAid": False, "fundedADA": False},
    "1819": {"basicAid": False, "fundedADA": False},
    "1920": {"basicAid": False, "fundedADA": True},
}
LCFF_UNPUBLISHED_REASON = {
    "1617": "CDE publishes no LCFF funding summary for this fiscal year.",
    "1718": "CDE publishes no LCFF funding summary for this fiscal year.",
    "1819": "CDE publishes no LCFF funding summary for this fiscal year.",
    "1920": "CDE's LCFF funding summary for this fiscal year does not "
            "publish a local revenue column, and basic-aid status is "
            "local revenue measured against the LCFF entitlement.",
}


def lcff_status(yy, fact, compute):
    """A fact the LCFF summary does not publish is said so, never guessed.

    `compute` is only called when the source publishes the fact, so a
    missing workbook can never fall through to a default."""
    if not LCFF_PUBLISHES.get(yy, {}).get(fact, True):
        return "not-published"
    return compute()


# The SACS database changed two names between FY2021-22 and FY2022-23.
# Verified by inspection on every year FY2016-17..FY2021-22, and by the
# shipped gate on the current years: the running pipeline reads
# Charters.SchoolCode and counts 58 county offices by the long spelling,
# so a mismatch there would already have failed.
SACS_ERA = {
    "legacy":  {"charterSchool": "SchoolID",   "coeType": "CO OFFICE"},
    "current": {"charterSchool": "SchoolCode", "coeType": "County Office"},
}
SACS_LEGACY_YEARS = {"1617", "1718", "1819", "1920", "2021", "2122"}


def vintage(yy):
    """The declared vocabulary for one source year, or a refusal."""
    if yy not in CE_VINTAGE:
        raise SystemExit(
            f"FY {yy}: no declared source vocabulary. CDE renames sheets and "
            "columns between vintages, and a year whose spellings have not "
            "been read off the published file cannot be parsed safely — the "
            "FY2018-19 vintage passed every gate on zero districts that way. "
            "Add it to CE_VINTAGE deliberately; nothing written.")
    sheet, co, dcode, dname = CE_VINTAGE[yy]
    era = SACS_ERA["legacy" if yy in SACS_LEGACY_YEARS else "current"]
    return {"sheet": sheet, "co": co, "dcode": dcode, "dname": dname, **era}


# THE WINDOW. Extended to nine years; SACS and the Current Expense
# workbook are published for every one of them (verified by fetch, on
# PK magic bytes rather than status). The LCFF summary is NOT — see
# LCFF_PUBLISHES, where the unknown set differs per FACT, not per year.
YEARS = [("1617", "2016-17"), ("1718", "2017-18"), ("1819", "2018-19"),
         ("1920", "2019-20"), ("2021", "2020-21"), ("2122", "2021-22"),
         ("2223", "2022-23"), ("2324", "2023-24"), ("2425", "2024-25")]
GATE_TOL = 0.05          # per-district vs published EDP 365
ROLLUP_TOL = 0.05        # per fund×object cell vs CDE's own totals

COUNTIES = [  # CDE county codes are alphabetical, 01..58
    "Alameda", "Alpine", "Amador", "Butte", "Calaveras", "Colusa",
    "Contra Costa", "Del Norte", "El Dorado", "Fresno", "Glenn",
    "Humboldt", "Imperial", "Inyo", "Kern", "Kings", "Lake", "Lassen",
    "Los Angeles", "Madera", "Marin", "Mariposa", "Mendocino", "Merced",
    "Modoc", "Mono", "Monterey", "Napa", "Nevada", "Orange", "Placer",
    "Plumas", "Riverside", "Sacramento", "San Benito", "San Bernardino",
    "San Diego", "San Francisco", "San Joaquin", "San Luis Obispo",
    "San Mateo", "Santa Barbara", "Santa Clara", "Santa Cruz", "Shasta",
    "Sierra", "Siskiyou", "Solano", "Sonoma", "Stanislaus", "Sutter",
    "Tehama", "Trinity", "Tulare", "Tuolumne", "Ventura", "Yolo", "Yuba"]

FN_GROUPS = [
    ("instruction",   "Instruction",                    1000, 1999),
    ("instrRelated",  "Instruction-related services",   2000, 2999),
    ("pupilServices", "Pupil services",                 3000, 3999),
    ("ancillary",     "Ancillary services",             4000, 4999),
    ("community",     "Community services",             5000, 5999),
    ("enterprise",    "Enterprise",                     6000, 6999),
    ("genAdmin",      "General administration",         7000, 7999),
    ("plant",         "Plant services",                 8000, 8999),
    ("otherOutgo",    "Other outgo",                    9000, 9999),
]
OBJ_FAMILIES = [
    ("certSalaries",  "Certificated salaries",  1000, 1999),
    ("classSalaries", "Classified salaries",    2000, 2999),
    ("benefits",      "Employee benefits",      3000, 3999),
    ("supplies",      "Books & supplies",       4000, 4999),
    ("services",      "Services & operations",  5000, 5999),
    ("capital",       "Capital outlay",         6000, 6999),
    ("otherOutgo",    "Other outgo",            7000, 7999),
]


def download(url, dest):
    if dest.exists():
        return
    print(f"  downloading {url}", file=sys.stderr)
    req = urllib.request.Request(url, headers={"User-Agent": "ca-ledger-pipeline/1.0"})
    # through the guard (#60): a fetched source lands read-only, so a test
    # cannot overwrite it afterwards
    cache_guard.write_cached(dest, urllib.request.urlopen(req, timeout=600).read(),
                             binary=True)


def ensure_year_files(yy, fy_dashed):
    CACHE.mkdir(parents=True, exist_ok=True)
    files = {
        "sacs_exe": (f"https://www3.cde.ca.gov/fiscal-downloads/sacs_data/"
                     f"{fy_dashed}/sacs{yy}.exe", CACHE / f"sacs{yy}.exe"),
        "alt_exe":  (f"https://www3.cde.ca.gov/fiscal-downloads/charterdata/"
                     f"{fy_dashed}/alt{yy}data.exe", CACHE / f"alt{yy}data.exe"),
        "ce":   (f"https://www.cde.ca.gov/ds/fd/ec/documents/currentexpense{yy}.xlsx",
                 CACHE / f"currentexpense{yy}.xlsx"),
        "lcff": (f"https://www.cde.ca.gov/fg/aa/pa/documents/lcffsummary{yy}.xlsx",
                 CACHE / f"lcffsummary{yy}.xlsx"),
    }
    download(*files["ce"])
    # A YEAR DECLARED TO HAVE NO LCFF WORKBOOK IS NOT FETCHED. Asking for
    # it would 404 and stop the build; the absence is already recorded in
    # LCFF_PUBLISHES and travels to the payload as not-published.
    if any(LCFF_PUBLISHES.get(yy, {}).values()) or yy not in LCFF_PUBLISHES:
        download(*files["lcff"])
    for key, mdb_name in (("sacs_exe", f"sacs{yy}.mdb"), ("alt_exe", f"alt{yy}data.mdb")):
        mdb = CACHE / mdb_name
        if not mdb.exists():
            url, exe = files[key]
            download(url, exe)
            with zipfile.ZipFile(exe) as z:   # the .exe is ZIP-compatible
                inner = [n for n in z.namelist() if n.lower().endswith(".mdb")][0]
                cache_guard.write_cached(mdb, z.read(inner), binary=True)
    return CACHE / f"sacs{yy}.mdb", CACHE / f"alt{yy}data.mdb", files["ce"][1], files["lcff"][1]


class StrictRow(dict):
    """A source row that REFUSES an absent column instead of returning None.

    THREE CONFIDENT WRONG NUMBERS in this sequence came from reading a
    column by a name the source does not use, where the miss was silent:

      CharterNum  vs CharterNumber   the charter qualifier fell back to
                                     the school code and under-counted
                                     the collisions 40 -> 32 (#65)
      LEAName     vs Dname           reported nine phantom key
                                     collisions, all with empty names (#63)
      SchoolCode  vs SchoolID        the Alternative Form's vintage
                                     rename, which at least raised (#61)

    An absent column and an empty column are indistinguishable through
    `.get()`, and the empty one produces an answer rather than an error.
    So the miss is made loud: `row["Nope"]` raises KeyError as a plain
    dict does, and `row.get("Nope")` — the form that used to be silent —
    raises too. A caller that genuinely wants "absent is acceptable"
    must say so with `row.optional("Nope")`, which is greppable and
    rare.

    This does not protect against a column that exists and holds the
    wrong thing. It protects against the failure mode that has actually
    occurred three times: a name that is not there at all.
    """

    __slots__ = ("_table",)

    def __init__(self, mapping, table):
        super().__init__(mapping)
        self._table = table

    def _missing(self, key):
        return KeyError(
            f"COLUMN {key!r} DOES NOT EXIST in {self._table}. Its columns "
            f"are: {', '.join(sorted(k for k in self if k))}. An absent "
            "column reads as empty and produces a confident wrong answer; "
            "use the real name, or row.optional() if absence is expected.")

    def __getitem__(self, key):
        if key not in self:
            raise self._missing(key)
        return super().__getitem__(key)

    def get(self, key, default=None):
        """Deliberately NOT dict.get: a typo must not read as an empty
        value. Use optional() where absence is a real possibility."""
        if key not in self:
            raise self._missing(key)
        return super().__getitem__(key)

    def optional(self, key, default=None):
        """The declared escape hatch: this column may legitimately be
        absent in some vintage, and the caller has decided what that
        means."""
        return super().get(key, default)


def mdb_rows(mdb, table):
    proc = subprocess.Popen(["mdb-export", str(mdb), table],
                            stdout=subprocess.PIPE, text=True)
    for row in csv.DictReader(proc.stdout):
        yield StrictRow(row, f"{Path(mdb).name}:{table}")
    proc.stdout.close()
    if proc.wait() != 0:
        raise SystemExit(f"mdb-export failed for {mdb}:{table}")


def slugify(s):
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", s.lower())).strip("-")


def assign_slugs(records, label, force_qualify=frozenset()):
    """Deterministic, collision-free slugs — the identifier contract.

    `records` is a list of (identity, name, qualifier) where `identity`
    is the source's own stable code (CDS for districts and county
    offices, the charter key for charters). The slug is a function of
    the published name alone when that name is unique, and of the name
    plus its qualifier when it is not — so no entity ever wins a bare
    slug that another entity of the same name could have won instead.

    Two properties this guarantees, both asserted in the test suite:

      1. The result depends only on the source data, never on the order
         Python happened to iterate a set. The old code sorted a set of
         code tuples keyed on the NAME alone, so for the duplicated
         names the tie was broken by set-iteration order — i.e. by
         PYTHONHASHSEED, which Python randomizes per process. The same
         source produced a different school-data.js, and therefore a
         different published SHA-256, on every run.
      2. Every name that is shared is disambiguated for EVERY holder of
         it. Letting one of three "Jefferson Elementary" districts hold
         the unqualified slug is a claim we cannot support: it names one
         district with an identifier that describes three.

    Raises rather than guessing if a qualifier fails to disambiguate.

    Returns (slug_by_identity, ambiguous), where `ambiguous` maps each
    unqualified slug that is NOT issued to the qualified slugs that
    share that name. Earlier builds handed the unqualified slug to an
    arbitrary one of them, so links using it are genuinely ambiguous —
    that map lets the page say which entities a stale link could have
    meant instead of silently picking one.
    """
    by_name = defaultdict(list)
    for ident, name, qualifier in records:
        by_name[slugify(name)].append((ident, name, qualifier))
    out, ambiguous = {}, {}
    for base in sorted(by_name):
        group = sorted(by_name[base], key=lambda r: r[0])
        # `force_qualify` names identities that shared this name in SOME
        # year even if they do not share it in the year that produced
        # this record. They must be qualified anyway: an unqualified
        # slug would describe two entities for that earlier year.
        if len(group) == 1 and group[0][0] not in force_qualify:
            out[group[0][0]] = base
            continue
        for ident, name, qualifier in group:
            out[ident] = slugify(name + "-" + str(qualifier))
        if len(group) > 1:
            ambiguous[base] = sorted(out[ident] for ident, _, _ in group)
    if len(set(out.values())) != len(out):
        dupes = sorted(s for s in set(out.values())
                       if list(out.values()).count(s) > 1)
        raise SystemExit(
            f"SLUG COLLISION in {label} after qualification: {dupes} — "
            "the qualifier does not disambiguate these records. "
            "Nothing written; a human must choose a stable identifier.")
    return out, ambiguous


def fn_group(code):
    if code.isdigit():
        n = int(code)
        for key, _, lo, hi in FN_GROUPS:
            if lo <= n <= hi:
                return key
    return "otherOutgo"


def obj_family(n):
    for key, _, lo, hi in OBJ_FAMILIES:
        if lo <= n <= hi:
            return key
    return "otherOutgo"


# CSAM Procedure 310 assigns the funding source of a resource ONLY by
# number range (there is no per-code source column in any CDE artifact).
# Ranges, verbatim: 0000-1999 unrestricted; 2000-9999 restricted, of
# which 3000-5999 federal, 6000-7999 state, 8000-9999 local. CDE assigns
# NO source to 2000-2999 (it sits before the federal block); we surface
# that range as its own group and never file it under a source. This
# grouping is the restricted/unrestricted split, finer: "unrestricted"
# here equals resource < 2000, exactly as edp_restr splits it — so the
# resource partition can never disagree with the displayed split.
RES_GROUPS = [
    ("unrestricted",    "Unrestricted"),
    ("federal",         "Federal restricted"),
    ("state",           "State restricted"),
    ("local",           "Local restricted"),
    ("restrictedOther", "Restricted — other (CDE assigns no source range)"),
]
STRS_ON_BEHALF_RES = "7690"   # On-Behalf Pension Contributions (state pays
# CalSTRS directly on the district's behalf; booked as object 3101 within
# resource 7690, entirely inside Current Expense — the district never
# touches this cash). Always shown as its own named row so a group total
# is never read as district-controlled spending. See face note below.
ELOP_RES = "2600"             # Expanded Learning Opportunities Program —
# the one live code in the unassigned 2000-2999 range; shown by name.
NAMED_FLOOR = 1_000_000.0     # named rows at >= $1M; the rest combine into
# an exact-remainder tail per group (the finding's hybrid).


def res_group(code):
    """Resource -> CSAM funding-source group. Non-digit codes fold into
    unrestricted, matching edp_restr's `isdigit() and >= 2000` rule, so
    the two views are one partition (empirically no alpha codes appear in
    Current Expense in any shipped vintage)."""
    if not code.isdigit():
        return "unrestricted"
    n = int(code)
    if n < 2000:
        return "unrestricted"
    if n < 3000:
        return "restrictedOther"
    if n < 6000:
        return "federal"
    if n < 8000:
        return "state"
    return "local"


OBJ_ORDER = [k for k, _, _, _ in OBJ_FAMILIES]


def obj_cells(res_obj):
    """V10a reduced cross-tab: the object-family split of ONE resource, as
    a fixed OBJ_FAMILIES-ordered whole-dollar array that sums EXACTLY to
    the resource's whole-dollar display value. The largest cell absorbs
    the rounding remainder (the same exact-remainder trick the group tails
    use), so the object breakdown ties to the resource total to the cent
    and the resource value is unchanged from V9. Trailing zeros trimmed.
    Returns (display_value, array)."""
    cents = {k: round(res_obj.get(k, 0.0), 2) for k in OBJ_ORDER}
    dv = round(round(sum(cents.values()), 2))     # == V9's round(resource$)
    dollars = {k: round(cents[k]) for k in OBJ_ORDER}
    diff = dv - sum(dollars.values())
    if diff:
        big = max(OBJ_ORDER, key=lambda k: (abs(dollars[k]), k))
        dollars[big] += diff
    arr = [dollars[k] for k in OBJ_ORDER]
    while arr and arr[-1] == 0:
        arr.pop()
    return dv, arr


def build_by_resource(res_map, res_obj_map=None):
    """The finding's hybrid, per district-year: CSAM-range groups, named
    rows for resources >= $1M (STRS on-behalf 7690 always named so it is
    never buried in a group total), and one exact-remainder tail per
    group. Named rows carry whole dollars (funding sources render as
    $Xm / per-ADA, never to the cent); the tail absorbs the remainder so
    every group STILL sums to the cent — group total = sum(named) + tail,
    and sum(group totals) = the gated Current Expense, exactly.
    When res_obj_map is given (V10a), each named row also carries its
    object-family split [code, whole$, [obj array]] — the reduced
    object × resource cross-tab, summing to the row's own total.
    Returns {group: {total (cents), named:[[code, whole$, obj?]], tail}}."""
    grouped = defaultdict(list)          # group -> [(code, value)]
    for code, v in res_map.items():
        grouped[res_group(code)].append((code, round(v, 2)))
    out = {}
    for gkey, _ in RES_GROUPS:
        rows = grouped.get(gkey, [])
        total = round(sum(v for _, v in rows), 2)
        named, named_sum = [], 0.0
        for code, v in sorted(rows, key=lambda cv: (-abs(cv[1]), cv[0])):
            if abs(v) >= NAMED_FLOOR or code == STRS_ON_BEHALF_RES:
                dv = round(v)            # whole dollars for display rows
                row = [code, dv]
                if res_obj_map is not None:
                    odv, arr = obj_cells(res_obj_map.get(code, {}))
                    row = [code, odv, arr]   # odv == dv by construction
                named.append(row)
                named_sum += dv
        out[gkey] = {"total": total, "named": named,
                     "tail": round(total - named_sum, 2)}
    return out


def slim_by_resource(by, named):
    """Ship the group total (`v`) for every non-empty group, every year —
    that is the funding-source answer, gated to the cent and nesting into
    restricted/unrestricted. The named-code rows (`n`) and their
    exact-remainder tail (`t`) are shipped for the latest year only: three
    years of every $1M+ code would be ~+18% on the file, so the finding's
    hybrid is held to the payload discipline by carrying the named detail
    on the current year and the group totals on all years. The pipeline
    still GATES the full breakout at full fidelity in every year."""
    out = {}
    for gkey, d in by.items():
        if abs(d["total"]) < 0.005 and not d["named"]:
            continue
        o = {"v": d["total"]}
        if named:
            if d["named"]:
                o["n"] = d["named"]
            if abs(d["tail"]) >= 0.005:
                o["t"] = d["tail"]
        out[gkey] = o
    return out


def edp_in_scope(obj, goal, func):
    """CDE's Current Expense of Education (EDP 365) row filter."""
    in_objects = (1000 <= obj <= 5999) or obj == 6500 or (7300 <= obj <= 7399)
    if not in_objects:
        return False
    gnum = int(goal) if goal.isdigit() else -1
    if 7100 <= gnum <= 7199 or goal == "8100":
        return False
    if func in ("3700", "8500"):
        return False
    if obj in (3701, 3702):
        return False
    return True


REV_GROUPS = (  # (key, ranges) — revenue objects, per the V7 finding
    ("lcffStateAid", [(8011, 8011), (8012, 8012), (8015, 8015), (8019, 8019)]),
    ("propertyTaxes", [(8020, 8089)]),
    ("lcffTransfers", [(8090, 8099)]),
    ("federal", [(8100, 8299)]),
    ("otherState", [(8300, 8599)]),
    ("localOther", [(8600, 8799)]),
)


def rev_group(n):
    for key, ranges in REV_GROUPS:
        for lo, hi in ranges:
            if lo <= n <= hi:
                return key
    return None


def process_year(yy, fy):
    V = vintage(yy)          # the declared vocabulary for THIS source year
    sacs, alt, ce_path, lcff_path = ensure_year_files(yy, fy)
    _alt_col = alt_school_col(yy)   # declared per vintage, never sniffed
    import openpyxl

    # ---- LEA + charter registries
    leas = {}
    for r in mdb_rows(sacs, "LEAs"):
        leas[(r["Ccode"], r["Dcode"])] = {
            "name": r["Dname"].strip(), "type": r["Dtype"].strip(),
            "ada": float(r["K12ADA"] or 0)}
    charters = {}
    for r in mdb_rows(sacs, "Charters"):
        charters[(r["Ccode"], r["Dcode"], r[V["charterSchool"]])] = {
            "name": r["CharterName"].strip(),
            "number": r["CharterNumber"].strip(),
            "level": r["ReportLevel"].strip(), "fund": (r.get("FundUsed") or "").strip(),
            "type": (r.get("ReportType") or "").strip(),
            "ada": float(r["K12ADA"] or 0)}
    n_coe = sum(1 for v in leas.values() if v["type"].startswith(V["coeType"]))
    if n_coe != 58:
        raise SystemExit(f"FY {fy}: {n_coe} COEs (expected 58) — nothing written")

    # ---- resource titles: CDE's OWN names, from THIS year's table only
    # (titles drift across vintages — 46 retitles in our window — so they
    # are never crosswalked across years; the FY2017 vintage lesson).
    res_title = {}
    for r in mdb_rows(sacs, "Resource"):
        res_title[r["Code"].strip()] = r["Title"].strip()

    # ---- stream the general ledger once
    edp = defaultdict(float)                 # (c,d) -> EDP365
    edp_fn = defaultdict(lambda: defaultdict(float))
    edp_fn_obj = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    edp_restr = defaultdict(lambda: [0.0, 0.0])   # [unrestricted, restricted]
    edp_by_res = defaultdict(lambda: defaultdict(float))    # (c,d) -> code -> $
    edp_res_obj = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    edp_by_obj = defaultdict(lambda: defaultdict(float))    # (c,d) -> objfam -> $
    edp_fn_grp = defaultdict(lambda: defaultdict(float))    # (c,d) -> (fn,grp) -> $
    res_stats = defaultdict(float)   # statewide: strsOnBehalf, indirect_<grp>
    coe_fn = defaultdict(lambda: defaultdict(float))
    coe_tot = defaultdict(float)
    charter_gl = defaultdict(lambda: defaultdict(float))   # (c,d,s) -> family
    charter_gl_tot = defaultdict(float)
    dep_funds_exp = defaultdict(float)       # (c,d) -> fund 09/62 exp
    rev = defaultdict(float)
    rev_detail = defaultdict(float)          # epa 8012, strs 8590, lottery 8560, passthrough
    rollup = defaultdict(float)              # ("99" or ccode, fund, object) -> value
    for r in mdb_rows(sacs, "UserGL"):
        c, d, s = r["Ccode"], r["Dcode"], r["SchoolCode"]  # SACS UserGL: stable across vintages, measured FY1617 and FY2223
        obj_s, fund = r["Object"], r["Fund"]
        v = float(r["Value"] or 0)
        rollup[("99", fund, obj_s)] += v
        rollup[(c, fund, obj_s)] += v
        obj = int(obj_s) if obj_s.isdigit() else -1
        if obj < 0:
            continue
        lea = leas.get((c, d))
        is_coe = lea and lea["type"].startswith("County Office")
        if 8000 <= obj <= 8799:
            g = rev_group(obj)
            if g:
                rev[g] += v
            if obj == 8012: rev_detail["epa"] += v
            # STRS on-behalf is object 8590 within resource 7690 only —
            # bare 8590 is the whole "all other state revenue" bucket
            if obj == 8590 and r["Resource"] == "7690":
                rev_detail["strsOnBehalf"] += v
            if obj == 8560: rev_detail["lottery"] += v
            if obj in (8287, 8587, 8697) or 8781 <= obj <= 8799:
                rev_detail["passThrough"] += v
            continue
        if not (1000 <= obj <= 9999):
            continue
        if s != "0000000":
            if 1000 <= obj <= 7999:
                charter_gl[(c, d, s)][obj_family(obj)] += v
                charter_gl_tot[(c, d, s)] += v
            continue
        if fund == "01":
            if edp_in_scope(obj, r["Goal"], r["Function"]):
                edp[(c, d)] += v
                g = fn_group(r["Function"])
                edp_fn[(c, d)][g] += v
                # V8 depth: object family within the gated scope, and the
                # restricted/unrestricted split (resource >= 2000 is
                # restricted, SACS's own definition)
                edp_fn_obj[(c, d)][g][obj_family(obj)] += v
                res = r["Resource"].strip()
                restricted = res.isdigit() and int(res) >= 2000
                edp_restr[(c, d)][1 if restricted else 0] += v
                # V9: funding source. Resource dollars per district, and
                # the function × resource-group cross-tab that must
                # reproduce every function total.
                edp_by_res[(c, d)][res] += v
                # V10a: the reduced object × resource cross-tab — object
                # family within each resource, and the plain object-family
                # margin the aggregate must reconcile to.
                of = obj_family(obj)
                edp_res_obj[(c, d)][res][of] += v
                edp_by_obj[(c, d)][of] += v
                edp_fn_grp[(c, d)][(g, res_group(res))] += v
                if res == STRS_ON_BEHALF_RES:
                    res_stats["strsOnBehalf"] += v
                if 7300 <= obj <= 7399:      # indirect-cost transfers
                    res_stats["indirect_" + res_group(res)] += v
            if is_coe and 1000 <= obj <= 7999:
                coe_fn[(c, d)][fn_group(r["Function"])] += v
                coe_tot[(c, d)] += v
        elif fund in ("09", "62") and 1000 <= obj <= 7999:
            dep_funds_exp[(c, d)] += v

    # ---- GATE 2: CDE's own rollups, cell by cell
    published = defaultdict(float)
    for r in mdb_rows(sacs, "UserGL_Totals"):
        c, d = r["Ccode"], r["Dcode"]
        scope = "99" if (c == "99" or d == "99999") else c
        published[(scope, r["Fund"], r["Object"])] += float(r["Value"] or 0)
    bad_cells = 0
    for key, pv in published.items():
        if abs(rollup.get(key, 0.0) - pv) > ROLLUP_TOL:
            bad_cells += 1
    if bad_cells:
        raise SystemExit(f"FY {fy}: {bad_cells} rollup cells disagree with "
                         "CDE's UserGL_Totals — nothing written")

    # ---- Alternative Form charters (+ their own totals gate)
    alt_data = defaultdict(lambda: defaultdict(float))
    alt_tot = defaultdict(float)
    alt_by_obj = defaultdict(float)
    for r in mdb_rows(alt, "Alternate_Form_Data"):
        o = r["ObjectCode"]
        v = float(r["total"] or 0)
        alt_by_obj[o] += v
        n = int(o) if o.isdigit() else -1
        if 1000 <= n <= 7999:
            alt_data[(r["Ccode"], r["Dcode"], r[_alt_col])][obj_family(n)] += v
            alt_tot[(r["Ccode"], r["Dcode"], r[_alt_col])] += v
        elif 8000 <= n <= 8799:
            g = rev_group(n)
            if g:
                rev[g] += v
            if n == 8012: rev_detail["epa"] += v
            if n == 8560: rev_detail["lottery"] += v
    bad_alt = 0
    for r in mdb_rows(alt, "Alternate_Form_Totals"):
        if abs(alt_by_obj.get(r["ObjectCode"], 0.0) - float(r["total"] or 0)) > ROLLUP_TOL:
            bad_alt += 1
    if bad_alt:
        raise SystemExit(f"FY {fy}: {bad_alt} Alternative Form totals disagree "
                         "— nothing written")

    # ---- GATE 1: published Current Expense of Education, to the cent
    wb = openpyxl.load_workbook(ce_path, read_only=True)
    if V["sheet"] not in wb.sheetnames:
        raise SystemExit(
            f"FY {fy}: the Current Expense workbook has no {V['sheet']!r} "
            f"sheet (it has {wb.sheetnames}). The declared vocabulary no "
            "longer matches the published file; nothing written.")
    ws = wb[V["sheet"]]
    rows = ws.iter_rows(values_only=True)
    header = None
    ce = {}
    for row in rows:
        if header is None:
            if row and str(row[0]).strip() == V["co"]:
                header = [str(x or "").replace("\n", " ").strip() for x in row]
            continue
        if row[0] is None:
            continue
        rec = dict(zip(header, row))
        c = str(rec[V["co"]]).zfill(2)
        d = str(rec[V["dcode"]]).zfill(5)
        ce[(c, d)] = {"name": str(rec[V["dname"]]).strip(),
                      "edp": float(rec["EDP 365"]),
                      "ada": float(rec["Current Expense ADA"]),
                      "type": str(rec["LEA Type"]).strip()}
    # THE GATE TARGET MUST EXIST. `ce` is the published control this year's
    # figures reconcile against, and it is built by locating a header row
    # whose first cell reads exactly "CO". A vintage that labels that column
    # "CO Code" leaves `header` None, `ce` EMPTY — and then every gate below
    # iterates nothing, accumulates no failures and reports success.
    # Measured on the FY2018-19 vintage: ce_rows=0, gate1_n=0, all green.
    # A year nobody verified would have shipped looking verified.
    gates.require_target(header,
                         "the Current Expense header row (first cell 'CO')",
                         "the published per-district control cannot be read.")
    gates.require_rows(len(ce), 900, "Current Expense districts",
                       "every per-district gate below would iterate nothing.")

    # THE CLASSIFICATION-SHAPE GATE (hard): per district, ADA>0 implies
    # instruction dollars > 0; statewide, the core function groups must
    # be nonzero. A function-code shift would keep the to-the-cent
    # totals gate green while gutting the classification.
    gates.require_rows(len(edp_fn), 900,
                       "districts with a function-coded ledger",
                       "the classification-shape gate would check nothing.")
    shape_fail = []
    sw = defaultdict(float)
    for key, fn in edp_fn.items():
        for g, v in fn.items():
            sw[g] += v
    for g in ("instruction", "instrRelated", "pupilServices", "genAdmin", "plant"):
        if sw.get(g, 0) <= 0:
            shape_fail.append(f"statewide {g!r} is zero")
    for (c, d), pub in ce.items():
        if pub["ada"] > 0 and edp_fn.get((c, d), {}).get("instruction", 0) <= 0:
            shape_fail.append(f"{pub['name']} ({c}-{d}): ADA {pub['ada']} "
                              "but zero instruction dollars")
    if shape_fail:
        for s in shape_fail[:10]:
            print("  SHAPE FAIL:", s, file=sys.stderr)
        raise SystemExit(f"FY {fy}: {len(shape_fail)} classification-shape "
                         "failure(s) — nothing written")

    gate_fail = []
    for (c, d), pub in ce.items():
        ours = edp.get((c, d), 0.0)
        if abs(ours - pub["edp"]) > GATE_TOL:
            gate_fail.append(f"{pub['name']} ({c}-{d}): computed ${ours:,.2f} "
                             f"vs published ${pub['edp']:,.2f}")
    if gate_fail:
        for g in gate_fail[:20]:
            print("  GATE FAIL:", g, file=sys.stderr)
        raise SystemExit(f"FY {fy}: {len(gate_fail)} district(s) fail the "
                         "Current Expense gate — nothing written")

    # ---- LCFF: basic aid + COE funded ADA
    #
    # A year CDE publishes no workbook for parses nothing here, and both
    # facts reach the payload through the not-published machinery. The
    # dicts stay empty rather than being filled with defaults — a default
    # is exactly the positive claim this layer refuses to make.
    lcff_absent = (yy in LCFF_PUBLISHES
                   and not any(LCFF_PUBLISHES[yy].values()))
    if lcff_absent:
        print(f"  FY {fy}: no LCFF workbook published — basic aid and funded "
              f"ADA are not-published for this year", file=sys.stderr)
    wb = None if lcff_absent else openpyxl.load_workbook(lcff_path,
                                                         read_only=True)
    sheet = None
    for name in ([] if wb is None else wb.sheetnames):
        n = name.replace(" ", "")
        if n.endswith("Annual") or n.endswith("AN"):
            sheet = name
    # A DECLARED ABSENCE IS NOT A PARSE FAILURE. The refusal still fires
    # for a year that IS supposed to have a workbook — that is the
    # vintage guard from #42 — but a year declared to publish nothing
    # skips the parse instead of stopping the build.
    if sheet is None and not lcff_absent:
        raise SystemExit(f"FY {fy}: no Annual sheet in {lcff_path.name}")
    ws = None if sheet is None else wb[sheet]
    header, basic_aid, coe_ada = None, {}, {}
    col = {}
    for row in ([] if ws is None else ws.iter_rows(values_only=True)):
        if header is None:
            if row and str(row[0]).strip() == "County Code":
                header = [str(x or "").strip() for x in row]
                for i, h in enumerate(header):
                    if h.startswith("Total LCFF Entitlement"): col["ent"] = i
                    elif "Local Revenue" in h: col["local"] = i
                    elif h.startswith("Total Funded ADA"): col["ada"] = i
                # HOW MANY COLUMNS THIS VINTAGE PUBLISHES IS DECLARED,
                # NOT SNIFFED. FY2019-20's workbook carries 12 columns
                # and no local-revenue column at all, where FY2020-21+
                # carry 20 including it. Requiring 3 everywhere would
                # refuse a legitimate vintage; requiring 2 everywhere
                # would let a broken parse through on a year that really
                # does publish local revenue. So the floor is per year,
                # taken from LCFF_PUBLISHES.
                need = 3 if LCFF_PUBLISHES.get(yy, {}).get("basicAid", True) else 2
                gates.require_rows(
                    len(col), need, f"FY {fy} LCFF header columns located",
                    f"header was {header}.")
            continue
        if row[0] is None:
            continue
        c, d, s = str(row[0]).zfill(2), str(row[1]).zfill(5), str(row[2]).zfill(7)
        def num(i):
            v = row[i]
            return float(v) if isinstance(v, (int, float)) else 0.0
        if s != "0000000":
            continue
        lea = leas.get((c, d))
        if not lea:
            continue
        if lea["type"].startswith("County Office"):
            coe_ada[(c, d)] = num(col["ada"])
        elif "local" in col:
            basic_aid[(c, d)] = num(col["local"]) > num(col["ent"])
        # else: this vintage publishes no local-revenue column, so basic
        # aid cannot be derived and is left unset — lcff_status() turns
        # that into "not-published" rather than a default False

    # ---- charter sponsorship / commingling per district
    sponsored = defaultdict(float)
    commingled = defaultdict(lambda: [0, 0.0])   # count, ada
    for (c, d, s), ch in charters.items():
        # ReportLevel values: CharterSchool / StateBoardOfEducation file
        # separately; SchoolDistrict / CountyOfficeOfEducation report inside
        # the authorizer. FundUsed is a word: "General" = Fund 01
        # (commingled), CharterSpecRevenue = Fund 09, CharterEnterprise = 62.
        if ch["level"] in ("CharterSchool", "StateBoardOfEducation"):
            sponsored[(c, d)] += ch["ada"]
        elif ch["level"] and ch["fund"] == "General":
            commingled[(c, d)][0] += 1
            commingled[(c, d)][1] += ch["ada"]

    # V8 PARENT-SUM GATES (hard): the depth partitions must equal the
    # gated Current Expense figure to the cent, district by district
    depth_fail = []
    for key, pub in ce.items():
        total = edp[key]
        fo = sum(v for fam in edp_fn_obj[key].values() for v in fam.values())
        ru = sum(edp_restr[key])
        if abs(fo - total) > 0.005 or abs(ru - total) > 0.005:
            depth_fail.append(f"{pub['name']}: fn×obj {fo:,.2f} / "
                              f"restr-split {ru:,.2f} vs {total:,.2f}")
    if depth_fail:
        for s in depth_fail[:8]:
            print("  V8 GATE FAIL:", s, file=sys.stderr)
        raise SystemExit(f"FY {fy}: {len(depth_fail)} depth partition(s) "
                         "do not sum to the gated figure — nothing written")

    # V9 RESOURCE GATES (hard, no write on failure): the funding-source
    # breakout must reproduce the gated figure to the cent, be a strict
    # refinement of the restricted/unrestricted split, and not distort
    # any function total.
    res_fail = []
    for key, pub in ce.items():
        total = round(edp[key], 2)
        by = build_by_resource(edp_by_res[key])
        # 1. group totals + the hybrid tails sum to Current Expense
        grp_sum = round(sum(g["total"] for g in by.values()), 2)
        disp_sum = round(sum(round(sum(v for _, v in g["named"]), 2) + g["tail"]
                             for g in by.values()), 2)
        if abs(grp_sum - total) > 0.005 or abs(disp_sum - total) > 0.005:
            res_fail.append(f"{pub['name']}: groups {grp_sum:,.2f} / "
                            f"displayed {disp_sum:,.2f} vs {total:,.2f}")
            continue
        # 2. resource partition == the displayed restricted/unrestricted
        unr = by["unrestricted"]["total"]
        restr = round(sum(by[g]["total"] for g, _ in RES_GROUPS
                          if g != "unrestricted"), 2)
        if (abs(unr - round(edp_restr[key][0], 2)) > 0.005 or
                abs(restr - round(edp_restr[key][1], 2)) > 0.005):
            res_fail.append(f"{pub['name']}: resource partition "
                            f"unr {unr:,.2f}/restr {restr:,.2f} != split "
                            f"{edp_restr[key][0]:,.2f}/{edp_restr[key][1]:,.2f}")
            continue
        # 3. function × resource-group reproduces every function total
        fn_from_grp = defaultdict(float)
        for (g, _grp), v in edp_fn_grp[key].items():
            fn_from_grp[g] += v
        for g, v in edp_fn[key].items():
            if abs(fn_from_grp[g] - v) > 0.005:
                res_fail.append(f"{pub['name']}: fn×group {g} "
                                f"{fn_from_grp[g]:,.2f} != {v:,.2f}")
                break
    if res_fail:
        for s in res_fail[:10]:
            print("  V9 GATE FAIL:", s, file=sys.stderr)
        raise SystemExit(f"FY {fy}: {len(res_fail)} resource partition(s) "
                         "fail the funding-source gates — nothing written")

    # V10a REDUCED CROSS-TAB GATES (hard, no write on failure): the
    # object × resource cross-tab must tie to BOTH margins — every
    # resource's object split sums to that resource's total, and the
    # object families aggregate across resources to the object-family
    # totals (V8), hence to Current Expense — and the SHIPPED whole-dollar
    # object arrays of the named resources must sum to each named value.
    xt_fail = []
    for key, pub in ce.items():
        ro = edp_res_obj[key]
        # margin 1: sum over object family within each resource == by_res
        for res, fams in ro.items():
            if abs(sum(fams.values()) - edp_by_res[key][res]) > 0.005:
                xt_fail.append(f"{pub['name']} res {res}: object split "
                               f"!= resource total"); break
        if xt_fail and xt_fail[-1].startswith(pub['name']):
            continue
        # margin 2: sum over resource within each object family == by_obj
        m_obj = defaultdict(float)
        for res, fams in ro.items():
            for of, v in fams.items():
                m_obj[of] += v
        bad_margin2 = any(abs(m_obj[of] - edp_by_obj[key][of]) > 0.005
                          for of in edp_by_obj[key])
        # and the whole cross-tab totals to Current Expense
        if bad_margin2 or abs(sum(m_obj.values()) - round(edp[key], 2)) > 0.05:
            xt_fail.append(f"{pub['name']}: cross-tab object margin or "
                           f"total off"); continue
        # the SHIPPED named object arrays sum to the named display value
        by = build_by_resource(edp_by_res[key], edp_res_obj[key])
        for g in by.values():
            for row in g["named"]:
                if len(row) == 3 and sum(row[2]) != row[1]:
                    xt_fail.append(f"{pub['name']} res {row[0]}: shipped "
                                   f"object array {sum(row[2])} != {row[1]}")
                    break
    if xt_fail:
        for s in xt_fail[:10]:
            print("  V10a GATE FAIL:", s, file=sys.stderr)
        raise SystemExit(f"FY {fy}: {len(xt_fail)} object×resource cross-tab(s) "
                         "fail the both-margins gate — nothing written")

    return {
        "leas": leas, "charters": charters, "ce": ce, "edp": edp,
        "edp_fn": edp_fn, "edp_fn_obj": edp_fn_obj, "edp_restr": edp_restr,
        "edp_by_res": edp_by_res, "edp_res_obj": edp_res_obj,
        "res_title": res_title,
        "res_stats": dict(res_stats), "coe_fn": coe_fn, "coe_tot": coe_tot,
        "coe_ada": coe_ada, "charter_gl": charter_gl,
        "charter_gl_tot": charter_gl_tot, "alt_data": alt_data,
        "alt_tot": alt_tot, "dep_funds_exp": dep_funds_exp,
        "basic_aid": basic_aid, "sponsored": sponsored,
        "commingled": commingled, "rev": dict(rev),
        "rev_detail": dict(rev_detail),
        "n_districts": len(ce),
    }


def main():
    ap = argparse.ArgumentParser(description="Rebuild school-data.js")
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()

    state_js = json.loads((ROOT / "data.js").read_text(encoding="utf-8")
                          .split("=", 1)[1].rsplit(";", 1)[0])

    years = {}
    for yy, fy in YEARS:
        print(f"FY {fy}:", file=sys.stderr)
        years[fy] = process_year(yy, fy)
        print(f"  gates passed: {years[fy]['n_districts']} districts to the "
              f"cent; rollups exact; alt totals exact", file=sys.stderr)

    latest = YEARS[-1][1]
    Y = [fy for _, fy in YEARS]
    YY_OF = {fy: yy for yy, fy in YEARS}

    # ---- NCES crosswalk (identifier matching for the address view)
    if not PUBSCHLS.exists():
        raise SystemExit(f"{PUBSCHLS} missing — download the CDE public "
                         "schools/districts directory export (see comment "
                         "at COMMON_ADMIN_NCES) before running")
    nces_ids = defaultdict(list)
    with open(PUBSCHLS, encoding="utf-8", errors="replace") as f:
        for r in csv.DictReader(f, delimiter="\t"):
            cds = (r.get("CDSCode") or "").strip()
            nces = (r.get("NCESDist") or "").strip()
            if len(cds) == 14 and cds.endswith("0000000") and nces and nces != "No Data":
                target = COMMON_ADMIN_NCES.get(nces, cds[:7])
                if nces not in nces_ids[target]:
                    nces_ids[target].append(nces)

    # ---- assemble districts (driven by the published CE list)
    districts = {}
    # titles for the codes that actually appear as named rows, per year —
    # CDE's own words, this vintage's table only; a code with no title in
    # its own year ships as its raw code (never invented, never grouped).
    res_titles_used = {fy: {} for fy in Y}
    # sorted on the CDS key itself — the source's own stable identity —
    # so the iteration order cannot depend on PYTHONHASHSEED.
    seen_keys = {k for fy in Y for k in years[fy]["ce"]}
    # A DECLARED RE-CODING MUST BE EXERCISED — both keys present, or the
    # declaration is a stale exemption.
    assert_recodings(seen_keys)
    # A re-coded district's OLD key folds into its new one, so the nine
    # years are one continuous record rather than two short ones. Every
    # year's figures are carried across; nothing is dropped.
    for _old, _rec in RECODED_DISTRICTS.items():
        _new = _rec["to"]
        for fy in Y:
            if _old in years[fy]["ce"] and _new not in years[fy]["ce"]:
                years[fy]["ce"][_new] = years[fy]["ce"].pop(_old)
            elif _old in years[fy]["ce"]:
                raise SystemExit(
                    f"RE-CODING COLLISION: {_old} and {_new} both present in "
                    f"FY {fy}. They cannot be the same district in a year "
                    "that contains both; nothing written.")
        # EVERY per-key table, DERIVED not listed. The first version of
        # this hand-wrote the table names and got one wrong — "by_res"
        # where the pipeline calls it "edp_by_res" — so Lowell Joint's
        # resource breakdown was silently left behind under the old key
        # and the resource gate refused. A hand-written list of names is
        # the same failure mode as a hand-written column name; this
        # finds every table that is keyed by (Ccode, Dcode) instead.
        for _d in (years[fy] for fy in Y):
            for _name, _t in list(_d.items()):
                if not isinstance(_t, dict) or _old not in _t:
                    continue
                if _new in _t:
                    raise SystemExit(
                        f"RE-CODING COLLISION in {_name}: {_old} and {_new} "
                        "are both present; nothing written.")
                _t[_new] = _t.pop(_old)
    all_keys = sorted({k for fy in Y for k in years[fy]["ce"]})
    newest_of = {k: next((years[fy]["ce"][k] for fy in reversed(Y)
                          if k in years[fy]["ce"]), None) for k in all_keys}
    district_slugs, district_ambig = assign_slugs(
        [(k, newest_of[k]["name"], COUNTIES[int(k[0]) - 1]) for k in all_keys],
        "K-12 districts")
    for key in all_keys:
        c, d = key
        newest = newest_of[key]
        slug = district_slugs[key]
        entry = {"name": newest["name"], "county": COUNTIES[int(c) - 1],
                 "type": newest["type"], "cds": c + d,
                 "nces": nces_ids.get(c + d, []), "years": {}}
        for fy in Y:
            yd = years[fy]
            if key not in yd["ce"]:
                continue
            pub = yd["ce"][key]
            fn = {k: round(v, 2) for k, v in yd["edp_fn"][key].items()}
            com = yd["commingled"].get(key)
            by_res = build_by_resource(yd["edp_by_res"][key],
                                       yd["edp_res_obj"][key])
            named_year = (fy == latest)
            if named_year:
                for g in by_res.values():
                    for row in g["named"]:
                        t = yd["res_title"].get(row[0])
                        if t:
                            res_titles_used[fy][row[0]] = t
            entry["years"][fy] = {
                "ada": round(pub["ada"], 2),
                "currentExpense": round(yd["edp"][key], 2),
                "cePublished": round(pub["edp"], 2),
                "byFunction": fn,
                "byFunctionObject": {g: {o: round(v, 2) for o, v in fam.items()}
                                     for g, fam in yd["edp_fn_obj"][key].items()},
                "byResource": slim_by_resource(by_res, named_year),
                "unrestricted": round(yd["edp_restr"][key][0], 2),
                "restricted": round(yd["edp_restr"][key][1], 2),
                # THREE-VALUED, AND NEVER A BARE BOOLEAN. `false` is a real
                # answer here, so a boolean has no room left to say "not
                # published" — .get(key, False) turned a missing LCFF file
                # into the positive claim that the district is state-funded.
                # The bare `basicAid` key is gone from EVERY year, so no
                # consumer can read it correctly in one year and be lied to
                # in another.
                # yy MUST come from THIS year, not the enclosing scope.
                # It was reading the outer loop's leftover `yy` — the
                # LAST year, which publishes everything — so all nine
                # years took the published branch and four of them
                # asserted "state-funded" about 3,756 districts on no
                # evidence. That is precisely the claim #55 exists to
                # prevent, reintroduced by a stale variable.
                "basicAidStatus": lcff_status(
                    YY_OF[fy], "basicAid",
                    lambda: ("basic-aid" if yd["basic_aid"].get(key)
                             else "state-funded")),
                "smallNSS": pub["ada"] < 250,
                "sponsoredCharterADA": round(yd["sponsored"].get(key, 0.0), 2),
                "commingledCharters": ({"count": com[0], "ada": round(com[1], 2)}
                                       if com else None),
                "depFundsCharterExp": round(yd["dep_funds_exp"].get(key, 0.0), 2),
            }
        districts[slug] = entry

    no_nces = [f"{e['name']} ({e['cds']})" for e in districts.values()
               if not e["nces"]]
    if no_nces:
        raise SystemExit("NCES ID MISSING for " + str(len(no_nces))
                         + " district(s) — the address view cannot match "
                         "them by identifier; nothing written:\n  "
                         + "\n  ".join(no_nces[:20]))

    # ---- county offices (records only)
    coes = {}
    coe_keys = sorted(k for k, lea in years[latest]["leas"].items()
                      if lea["type"].startswith("County Office"))
    coe_slugs, coe_ambig = assign_slugs(
        [(k, years[latest]["leas"][k]["name"], COUNTIES[int(k[0]) - 1])
         for k in coe_keys], "county offices")
    for key in coe_keys:
        lea = years[latest]["leas"][key]
        c, d = key
        entry = {"name": lea["name"], "county": COUNTIES[int(c) - 1],
                 "cds": c + d, "years": {}}
        for fy in Y:
            yd = years[fy]
            if key not in yd["coe_tot"]:
                continue
            # A NUMBER IS PUBLISHED ONLY WHEN IT IS KNOWN. Zero is a real
            # answer for funded ADA, so zero cannot carry "not published";
            # absence of the key yields NaN downstream, which is not a
            # valid answer and cannot be mistaken for one.
            entry["years"][fy] = {
                **({"fundedADA": round(yd["coe_ada"][key], 2)}
                   if LCFF_PUBLISHES.get(yy, {}).get("fundedADA", True)
                   and key in yd["coe_ada"] else {}),
                "expenditures": round(yd["coe_tot"][key], 2),
                "byFunction": {k: round(v, 2) for k, v in yd["coe_fn"][key].items()},
            }
        coes[coe_slugs[key]] = entry

    # ---- charters that file separately; dependent ones as pointers
    charters_out, dependents = {}, []
    latest_reg = years[latest]["charters"]
    # sorted on the charter key itself, for the same reason as districts.
    all_charter_keys = sorted(
        {k for fy in Y for k in list(years[fy]["charter_gl"]) + list(years[fy]["alt_data"])})
    charter_reg = {}
    for key in all_charter_keys:
        reg = next((years[fy]["charters"][k] for fy in reversed(Y)
                    for k in [key] if k in years[fy]["charters"]), None)
        if reg is not None:
            charter_reg[key] = reg
    # the charter number is NOT unique on its own (0756 is shared by the
    # nine High Tech schools), so it qualifies a shared NAME, never
    # identifies a charter; assign_slugs raises if that is not enough.
    # ── IDENTITY IS (COUNTY, SCHOOL CODE), NOT (NAME, CHARTER NUMBER) ──
    #
    # A charter's authorizer can change — to another district, or to a
    # direct-funded block — and the Dcode moves with it. Keying on
    # (Ccode, Dcode, SchoolCode) therefore split 33 charters into two
    # records each, and no qualifier could separate them, because they
    # were never two entities: assign_slugs refused, correctly.
    #
    # (county, school code) is the pair the source keeps stable across
    # every one of those transitions. Verified before this was written
    # (docs/V17A_CHARTER_REKEY_PREREQS.md, and the two checks in this
    # PR): it maps to 2+ rows within a single year ZERO times, in the
    # registry AND in both financial tables, across all nine years.
    #
    # THE QUALIFIER IS THE SCHOOL CODE, not the charter number. The
    # number is not unique — measured, ten pairs of distinct identities
    # share both a name and a number (Impact Academy of Arts &
    # Technology 0836 holds school codes 0113902 and 0137646), so it
    # leaves ten collisions standing. The school code is the identity's
    # own second half and leaves none.
    charter_identity = {}          # (c, d, s) -> (c, s)
    identity_years = defaultdict(dict)
    for k in all_charter_keys:
        if k in charter_reg:
            charter_identity[k] = (k[0], k[2])
    # THE IDENTITY'S NAME COMES FROM ITS NEWEST YEAR, exactly as
    # newest_of does for districts. Taking whichever key iterated last
    # gave two records the same name and different slugs: Pacific View
    # Charter appeared as both `pacific-view-charter` and
    # `pacific-view-charter-2-0`, because one identity's older key
    # carried the earlier name "Pacific View Charter 2.0".
    ident_newest_year = {}
    ident_reg = {}
    for k in all_charter_keys:
        if k not in charter_reg:
            continue
        cs = charter_identity[k]
        newest = max((i for i, fy in enumerate(Y)
                      if k in years[fy]["charters"]), default=-1)
        if newest > ident_newest_year.get(cs, -2):
            ident_newest_year[cs] = newest
            ident_reg[cs] = charter_reg[k]
    # A NAME SHARED IN *ANY* YEAR IS A SHARED NAME. Qualifying on the
    # newest name alone missed a real ambiguity: Pacific View Charter
    # (Humboldt 12/1230150) was renamed "Pacific View Charter 2.0" in
    # FY2017-18, but in FY2016-17 it carried the same name as the
    # distinct San Diego charter 37/3731221. Under newest-name-only
    # qualification the base names differ, so San Diego took the bare
    # `pacific-view-charter` — an identifier that, for FY2016-17,
    # describes two entities. They are two entities, not one: different
    # counties, different school codes, both present in FY2016-17.
    #
    # So every name an identity has EVER held counts toward ownership,
    # which is #22's rule applied across the window rather than at its
    # end: every holder of a shared name is disambiguated.
    ident_all_names = defaultdict(set)
    for k in all_charter_keys:
        if k not in charter_reg:
            continue
        cs = charter_identity[k]
        for fy in Y:
            reg_fy = years[fy]["charters"].get(k)
            if reg_fy and reg_fy.get("name"):
                ident_all_names[cs].add(slugify(reg_fy["name"]))
        ident_all_names[cs].add(slugify(charter_reg[k]["name"]))
    shared_bases = {b for b, owners in
                    ((b, [cs for cs, names in ident_all_names.items() if b in names])
                     for b in {n for names in ident_all_names.values() for n in names})
                    if len(owners) > 1}
    charter_slugs_by_ident, charter_ambig = assign_slugs(
        [(cs, r["name"], cs[1].lower()) for cs, r in ident_reg.items()],
        "charters", force_qualify={cs for cs, names in ident_all_names.items()
                                   if names & shared_bases})
    # A base name shared ACROSS YEARS is ambiguous for every identity
    # that ever held it, including those whose current name differs. The
    # list must name them all, or it says a stale link could have meant
    # one thing when it could have meant two.
    for base in shared_bases:
        owners = sorted(charter_slugs_by_ident[cs]
                        for cs, names in ident_all_names.items()
                        if base in names and cs in charter_slugs_by_ident)
        if len(owners) > 1:
            charter_ambig[base] = owners
    charter_slugs = {k: charter_slugs_by_ident[charter_identity[k]]
                     for k in charter_identity}
    # ── RETIRED SLUGS FROM THE OLD KEY (the #22 treatment) ────────────
    # Charters used to be slugged on (name, CHARTER NUMBER). Where the
    # number qualified a shared name, that slug is not reachable under
    # the new key, and a link carrying it is genuinely ambiguous — the
    # number identified more than one charter. assign_slugs already
    # records the names it qualifies, but not these, because the
    # QUALIFIER changed rather than the name. They are recorded here so
    # the page can say what a stale link could have meant instead of
    # redirecting to an arbitrary one of them.
    retired = defaultdict(set)
    for k, r in charter_reg.items():
        base = slugify(r["name"])
        old_slug = f"{base}-{(r['number'] or k[2]).lower()}"
        new_slug = charter_slugs[k]
        if old_slug != new_slug:
            retired[old_slug].add(new_slug)
    # A retired slug that maps to MORE THAN ONE record is genuinely
    # ambiguous — the old charter number identified several charters, so
    # the page must say which it could have meant rather than pick one.
    # A retired slug that maps to exactly one is an unambiguous rename,
    # which is a different statement and is recorded separately: saying
    # "this is now X" there is a fact, not a guess.
    charter_renamed = {}
    for old_slug, news in retired.items():
        if len(news) > 1:
            charter_ambig.setdefault(old_slug, [])
            for n in news:
                if n not in charter_ambig[old_slug]:
                    charter_ambig[old_slug].append(n)
        else:
            charter_renamed[old_slug] = next(iter(news))
    for v in charter_ambig.values():
        v.sort()
    for key in all_charter_keys:
        reg = charter_reg.get(key)
        if reg is None:
            continue
        c, d, s = key
        slug = charter_slugs[key]
        auth = years[latest]["leas"].get((c, d)) or years[Y[0]]["leas"].get((c, d))
        entry = charters_out.get(slug) or {
            "name": reg["name"], "county": COUNTIES[int(c) - 1],
            "charterNumber": reg["number"],
            "authorizer": auth["name"] if auth else "State Board of Education",
            "years": {}}
        for fy in Y:
            yd = years[fy]
            if key in yd["charter_gl"]:
                entry["years"][fy] = {
                    "mode": "sacs",
                    # THE AUTHORIZER IS A PER-YEAR FACT (V17a prereq 3):
                    # who oversaw the charter, and how its money reached
                    # it, in THIS year — not an attribute of the record.
                    "authorizer": (yd["leas"].get((c, d)) or {}).get(
                        "name", "State Board of Education"),
                    "ada": round(yd["charters"].get(key, {}).get("ada", 0.0), 2),
                    "expenditures": round(yd["charter_gl_tot"][key], 2),
                    "byObject": {k: round(v, 2) for k, v in yd["charter_gl"][key].items()},
                }
            elif key in yd["alt_data"]:
                entry["years"][fy] = {
                    "mode": "alt",
                    "authorizer": (yd["leas"].get((c, d)) or {}).get(
                        "name", "State Board of Education"),
                    "ada": round(yd["charters"].get(key, {}).get("ada", 0.0), 2),
                    "expenditures": round(yd["alt_tot"][key], 2),
                    "byObject": {k: round(v, 2) for k, v in yd["alt_data"][key].items()},
                }
        if entry["years"]:
            charters_out[slug] = entry
    # ── THE AUTHORIZER CHANGE IS A COMPARABILITY FACT (V17a prereq 3) ──
    # Derived from what the record now shows, not declared: after the
    # re-key the pipeline can see both sides of the change. A reader
    # comparing across it is comparing a period of district oversight
    # against a period under a different authorizer, or under direct
    # funding. Stated, never characterised — a charter moving to a
    # direct-funded block is routine under California charter law.
    for slug, entry in charters_out.items():
        seq = [(fy, entry["years"][fy].get("authorizer"))
               for fy in Y if fy in entry["years"]]
        changes = [(seq[i - 1][0], seq[i][0], seq[i - 1][1], seq[i][1])
                   for i in range(1, len(seq))
                   if seq[i][1] and seq[i - 1][1] and seq[i][1] != seq[i - 1][1]]
        if changes:
            entry["authorizerChanges"] = [
                {"from": a, "to": b, "fromName": x, "toName": y}
                for a, b, x, y in changes]
            entry["authorizerNote"] = (
                "The authorizer of this charter changed during the years "
                "shown: " + "; ".join(
                    f"{x} through FY {a}, then {y} from FY {b}"
                    for a, b, x, y in changes)
                + ". The authorizer oversees the charter and is the route "
                  "its funding takes, so figures either side of that "
                  "change are reported under different oversight.")
    FUND_LABEL = {"General": "Fund 01", "CharterSpecRevenue": "Fund 09",
                  "CharterEnterprise": "Fund 62"}
    # name, then the charter key — an explicit tie-break so the shipped
    # order of same-named dependents cannot drift between builds.
    for key, ch in sorted(latest_reg.items(), key=lambda kv: (kv[1]["name"], kv[0])):
        if ch["level"] in ("CharterSchool", "StateBoardOfEducation") or not ch["level"]:
            continue
        c, d, s = key
        auth = years[latest]["leas"].get((c, d))
        dependents.append({
            "name": ch["name"], "county": COUNTIES[int(c) - 1],
            "authorizer": auth["name"] if auth else "",
            "fund": FUND_LABEL.get(ch["fund"], ch["fund"]),
            "ada": round(ch["ada"], 2),
            "commingled": ch["fund"] == "General"})

    # ---- the overlap block, live
    agency = None
    for fy in Y:
        b = state_js["budgets"].get(fy)
        if not b:
            continue
        for a in b["agencies"]:
            if "K thru 12" in a["name"]:
                agency = agency or {}
                agency[fy] = round(a["gf"] + a["sp"] + a["bd"], 3)
    overlap_years = {}
    for fy in Y:
        yd = years[fy]
        rev, det = yd["rev"], yd["rev_detail"]
        total = sum(rev.values())
        state_sourced = rev["lcffStateAid"] + rev["otherState"] - det.get("lottery", 0)
        overlap_years[fy] = {
            "leaRevenuesB": round(total / 1e9, 3),
            "stateSourcedB": round(state_sourced / 1e9, 3),
            "stateShare": round(state_sourced / total, 4),
            "propertyTaxB": round(rev["propertyTaxes"] / 1e9, 3),
            "propertyTaxShare": round(rev["propertyTaxes"] / total, 4),
            "federalB": round(rev["federal"] / 1e9, 3),
            "federalShare": round(rev["federal"] / total, 4),
            "agencyEnactedB": agency.get(fy) if agency else None,
            "agreement": (round(state_sourced / 1e9 / agency[fy], 4)
                          if agency and agency.get(fy) else None),
            "epaB": round(det.get("epa", 0) / 1e9, 3),
            "strsOnBehalfB": round(det.get("strsOnBehalf", 0) / 1e9, 3),
            "interLeaPassThroughB": round(det.get("passThrough", 0) / 1e9, 3),
        }

    payload = {
        "meta": {
            "source": "cde.ca.gov",
            "sourceLabel": "California Department of Education — SACS unaudited "
                           "actuals, Current Expense of Education, and LCFF "
                           "summary data",
            "generated": date.today().isoformat(),
            "units": "as-reported dollars (exact); ADA as published",
            "identifiers": "Slugs are a function of the published name, "
                           "qualified by county (districts, county offices) "
                           "or charter number when a name is shared. Records "
                           "are keyed on CDE's own CDS code throughout, so "
                           "the same source data always produces the same "
                           "identifiers and the same digest.",
            # WHY a fact is unknown, keyed by (year, fact) — a property
            # of CDE's publication, not of each of 941 districts, so it is
            # stated once and looked up.
            #
            # AT META TOP LEVEL, which is where schools.html reads it.
            # It was nested inside meta.resource — 16 spaces of
            # indentation rather than 12 — so the payload carried the
            # reasons and the page could not see them: `D.meta.unpublished`
            # was undefined, unpubWhy() returned null, and the record said
            # "not known" with no reason attached. The status was right;
            # the explanation never arrived.
            "unpublished": {
                fy: {fact: LCFF_UNPUBLISHED_REASON[yy]
                     for fact, published
                     in LCFF_PUBLISHES.get(yy, {}).items() if not published}
                for yy, fy in YEARS
                if not all(LCFF_PUBLISHES.get(yy, {}).values())
            },
            "ambiguousSlugs": {k: v for k, v in (
                ("districts", district_ambig), ("countyOffices", coe_ambig),
                ("charters", charter_ambig)) if v},
            # a retired identifier that maps to exactly one record: not
            # ambiguous, so stated as a rename rather than a choice
            "renamedSlugs": {"charters": charter_renamed} if charter_renamed else {},
            "basis": "Unaudited actual expenditures as filed by each LEA under "
                     "SACS. District figures are the Current Expense of "
                     "Education (Fund 01 current operating expense, CDE's EDP "
                     "365) and REPRODUCE CDE'S PUBLISHED PER-DISTRICT FIGURE "
                     "TO THE CENT before publication; county-office and "
                     "charter figures reconcile through CDE's own published "
                     "county/state rollups and Alternative Form totals.",
            "gates": {
                "district": "recomputed EDP 365 == published Current Expense "
                            f"of Education within ${GATE_TOL:.2f}, every "
                            "district-year (932/932 to the cent, FY 2024-25)",
                "rollup": "sum(UserGL) == CDE's UserGL_Totals for every state "
                          "and county fund-by-object cell; Alternative Form "
                          "data == its totals table",
            },
            "comparisonScope": "School districts only, per ADA. County offices "
                               "of education are records only — CDE itself "
                               "excludes them from its per-ADA statistic. "
                               "Charters are records only. K-12 entities are "
                               "never compared to or summed with cities, "
                               "counties, special districts, or the state.",
            "overlap": {
                "years": overlap_years,
                "latest": latest,
                "statement": "computed on every pipeline run; the UI must "
                             "render these values and hardcode none",
                "traps": "EPA (Prop 30/55) is continuously appropriated "
                         "outside the annual Budget Act but inside the DOF "
                         "General Fund display; STRS on-behalf contributions "
                         "appear in both layers at different values; "
                         "statewide LEA sums contain inter-LEA pass-throughs; "
                         "Proposition 98 is a K-14 guarantee that also counts "
                         "property taxes — it is neither the agency total "
                         "nor total school funding.",
                "agreementNote": "enacted appropriations vs year-end accrual "
                                 "actuals agree to roughly 1-3.5 percent and "
                                 "never to the dollar",
            },
            "resource": {
                "groups": [{"key": k, "name": n,
                            "restricted": k != "unrestricted"}
                           for k, n in RES_GROUPS],
                "namedFloorM": round(NAMED_FLOOR / 1e6, 1),
                "namedYear": latest,   # named codes shipped for this year;
                # all years carry gated group totals (payload discipline)
                "strsOnBehalfRes": STRS_ON_BEHALF_RES,
                "elopRes": ELOP_RES,
                # live statewide figures per year (Current Expense scope,
                # districts) so the UI hardcodes no dollar amount
                "stats": {fy: {
                    "strsOnBehalfB": round(years[fy]["res_stats"]
                                           .get("strsOnBehalf", 0) / 1e9, 3),
                    "indirectToUnrestrictedB": round(-years[fy]["res_stats"]
                        .get("indirect_unrestricted", 0) / 1e9, 3),
                } for fy in Y},
                # the four face requirements, stated once, plainly
                "onBehalfNote": "Resource 7690, On-Behalf Pension "
                    "Contributions, is money the State pays to CalSTRS "
                    "directly on the district's behalf. It is inside the "
                    "Current Expense of Education but the district never "
                    "receives or spends this cash — it is not district "
                    "spending. Shown as its own row wherever it appears.",
                "indirectNote": "Per-source figures include each program's "
                    "share of district overhead, transferred into it at the "
                    "district's approved indirect-cost rate; unrestricted is "
                    "shown net of those reimbursements. The transfers net to "
                    "zero across the district, so totals reconcile exactly.",
                "unrestrictedNote": "Unrestricted is not local. Resource 0000 "
                    "and the rest of the unrestricted range hold LCFF "
                    "apportionment, the Education Protection Account, and "
                    "unrestricted Lottery — largely STATE money the district "
                    "may spend without a categorical restriction. Local "
                    "restricted is a separate group.",
                "lcffNote": "LCFF is accounted for as a single unrestricted "
                    "resource. Base, supplemental, and concentration grants "
                    "are NOT tracked separately in any district's general "
                    "ledger (CDE's LCFF FAQ; California State Auditor "
                    "2019-101), so no base-vs-supplemental breakout exists "
                    "here — the ledger does not contain one.",
                "otherRangeNote": "The 2000-2999 range has no funding source "
                    "assigned in CDE's classification (CSAM Procedure 310); "
                    "its codes are shown by their official names, never filed "
                    "under a source.",
                "objectSplitNote": "Each named funding source expands to what "
                    "it bought, by object — the same salary / benefit / "
                    "supply / service families the function view uses. The "
                    "objects sum to the funding source's total to the cent, "
                    "and across all sources to the same object totals. This "
                    "is the reduced cross-tab: a source's own object split, "
                    "never the full object-by-source grid.",
            },
            "commingledNote": "Charters reported inside an authorizer's Fund "
                              "01 cannot be separated from the district's own "
                              "figures; affected district records state the "
                              "count and ADA on their face.",
            "jpas": "ROPs, SELPAs, and other JPAs file to CDE and are inside "
                    "the rollup gates but are out of scope as records in this "
                    "version.",
        },
        "years": Y,
        "functions": [{"key": k, "name": n} for k, n, _, _ in FN_GROUPS],
        "objectFamilies": [{"key": k, "name": n} for k, n, _, _ in OBJ_FAMILIES],
        "resourceTitles": res_titles_used,
        "districts": districts,
        "countyOffices": coes,
        "charters": charters_out,
        "dependentCharters": dependents,
    }
    prev = revisions.previous_payload(OUT_PATH)
    stamp(payload)
    body = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    print(f"{len(districts)} districts · {len(coes)} county offices · "
          f"{len(charters_out)} charter records · {len(dependents)} dependent "
          f"charter pointers · payload ≈ {len(body) / 1048576:.2f} MB",
          file=sys.stderr)
    ol = overlap_years[latest]
    print(f"overlap FY {latest}: state-sourced {ol['stateShare']*100:.1f}% of "
          f"${ol['leaRevenuesB']:.1f}B LEA revenues; agency "
          f"${ol['agencyEnactedB']}B; agreement {ol['agreement']}",
          file=sys.stderr)
    if not args.write:
        print("Dry run — nothing written. Use --write.", file=sys.stderr)
        return
    header = ("/* GENERATED by pipeline/fetch_school_data.py on "
              f"{date.today().isoformat()} — do not edit by hand. */\n")
    OUT_PATH.write_text(header + "window.CA_SCHOOL_DATA = " + body + ";\n",
                        encoding="utf-8")

    revisions.record_revision('school', prev, payload)
    print(f"Wrote {OUT_PATH} ({OUT_PATH.stat().st_size / 1048576:.2f} MB)")


if __name__ == "__main__":
    main()
